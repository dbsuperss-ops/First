import sys
import openpyxl
import os

sys.stdout.reconfigure(encoding='utf-8')

base = r'C:\Users\dbsup\OneDrive - Kyungshin Corporation\02_경영_보고_전략\사업계획\2026'

def dump(filepath, sheetname, min_row=1, max_row=30, min_col=1, max_col=20):
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    if sheetname not in wb.sheetnames:
        print(f'  [시트 없음: {sheetname}]')
        wb.close()
        return
    ws = wb[sheetname]
    print(f'  [시트: {sheetname}]')
    for r in range(min_row, max_row + 1):
        row_vals = []
        for c in range(min_col, max_col + 1):
            v = ws.cell(row=r, column=c).value
            row_vals.append(str(v)[:14] if v is not None else '')
        if any(v.strip() for v in row_vals):
            print(f'  R{r:3d}: ' + ' | '.join(row_vals))
    wb.close()

# KCTR Summary rows 1~30 to find 매출액 and 매출원가 positions
print('=== KCTR Summary rows 1~30 ===')
dump(os.path.join(base, 'KCTR_2026년 사업계획.xlsx'), 'Summary', min_row=1, max_row=30, max_col=19)

print()
# KSCCZ 손익계획 rows 1~30
print('=== KSCCZ 손익계획 rows 1~30 ===')
dump(os.path.join(base, 'KSCCZ_2026년 사업계획.xlsx'), '손익계획', min_row=1, max_row=30, max_col=20)

print()
# KSCCZ 제조원가계획 rows 1~25
print('=== KSCCZ 제조원가계획 rows 1~25 ===')
dump(os.path.join(base, 'KSCCZ_2026년 사업계획.xlsx'), '제조원가계획', min_row=1, max_row=25, max_col=20)

print()
# KSCI 2026 Summary rows 1~40
print('=== KSCI 2026 Summary rows 1~40 ===')
dump(os.path.join(base, 'KSCI_2026년 사업계획.xlsx'), '2026 Summary', min_row=1, max_row=40, max_col=20)
