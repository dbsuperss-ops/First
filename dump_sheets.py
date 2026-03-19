import sys
import openpyxl
import os

sys.stdout.reconfigure(encoding='utf-8')

base = r'C:\Users\dbsup\OneDrive - Kyungshin Corporation\02_경영_보고_전략\사업계획\2026'

def dump_sheet(filepath, sheetname, max_row=35, max_col=20):
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    if sheetname not in wb.sheetnames:
        print(f'  [시트 없음: {sheetname}]')
        wb.close()
        return
    ws = wb[sheetname]
    print(f'  [시트: {sheetname}]')
    for r in range(1, max_row + 1):
        row_vals = []
        for c in range(1, max_col + 1):
            v = ws.cell(row=r, column=c).value
            row_vals.append(str(v)[:14] if v is not None else '')
        if any(v.strip() for v in row_vals):
            print(f'  R{r:3d}: ' + ' | '.join(row_vals))
    wb.close()

print('=== KSCCZ: 손익계획 ===')
dump_sheet(os.path.join(base, 'KSCCZ_2026년 사업계획.xlsx'), '손익계획')

print()
print('=== KSCCZ: 제조원가계획 ===')
dump_sheet(os.path.join(base, 'KSCCZ_2026년 사업계획.xlsx'), '제조원가계획')

print()
print('=== KSCP: 사업계획 ===')
dump_sheet(os.path.join(base, 'KSCP_2026년 사업계획.xlsx'), '사업계획', max_col=30)

print()
print('=== KSCE: PL2026(RS) ===')
dump_sheet(os.path.join(base, 'KSCE_2026년 사업계획.xlsx'), 'PL2026(RS)')

print()
print('=== KSCE: PL2026(KR) ===')
dump_sheet(os.path.join(base, 'KSCE_2026년 사업계획.xlsx'), 'PL2026(KR)')

print()
print('=== KSCI: 2026 Summary ===')
dump_sheet(os.path.join(base, 'KSCI_2026년 사업계획.xlsx'), '2026 Summary')

print()
print('=== KCTR: Summary ===')
dump_sheet(os.path.join(base, 'KCTR_2026년 사업계획.xlsx'), 'Summary')

print()
print('=== KCTR: 26 Detail ===')
dump_sheet(os.path.join(base, 'KCTR_2026년 사업계획.xlsx'), '26 Detail')
