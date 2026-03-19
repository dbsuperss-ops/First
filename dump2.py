import sys
import openpyxl
import os

sys.stdout.reconfigure(encoding='utf-8')

base = r'C:\Users\dbsup\OneDrive - Kyungshin Corporation\02_경영_보고_전략\사업계획\2026'

def dump_sheet(filepath, sheetname, max_row=10, min_col=1, max_col=40):
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    if sheetname not in wb.sheetnames:
        print(f'  [시트 없음: {sheetname}]')
        wb.close()
        return
    ws = wb[sheetname]
    print(f'  [시트: {sheetname}]')
    for r in range(1, max_row + 1):
        row_vals = []
        for c in range(min_col, max_col + 1):
            v = ws.cell(row=r, column=c).value
            row_vals.append(str(v)[:12] if v is not None else '')
        if any(v.strip() for v in row_vals):
            print(f'  R{r:3d}: ' + ' | '.join(row_vals))
    wb.close()

# KSCE: PL2026(RS) - col 1~40 확인, row 1~8만
print('=== KSCE PL2026(RS) 헤더 (col 1~40) ===')
dump_sheet(os.path.join(base, 'KSCE_2026년 사업계획.xlsx'), 'PL2026(RS)', max_row=8, max_col=40)

print()
# KSCE: PL2026(KR) 헤더
print('=== KSCE PL2026(KR) 헤더 (col 1~40) ===')
dump_sheet(os.path.join(base, 'KSCE_2026년 사업계획.xlsx'), 'PL2026(KR)', max_row=8, max_col=40)

print()
# KCTR: Summary 전체
print('=== KCTR Summary 전체 (row 30~80) ===')
wb = openpyxl.load_workbook(os.path.join(base, 'KCTR_2026년 사업계획.xlsx'), read_only=True, data_only=True)
ws = wb['Summary']
for r in range(30, 81):
    row_vals = [str(ws.cell(row=r, column=c).value or '')[:14] for c in range(1, 20)]
    if any(v.strip() for v in row_vals):
        print(f'  R{r:3d}: ' + ' | '.join(row_vals))
wb.close()

print()
# KSCI: Report 시트 확인
print('=== KSCI: Report 시트 (첫 50행) ===')
dump_sheet(os.path.join(base, 'KSCI_2026년 사업계획.xlsx'), 'Report', max_row=50, max_col=20)
