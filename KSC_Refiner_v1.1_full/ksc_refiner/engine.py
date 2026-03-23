import os
import sys
import json

sys.stdout.reconfigure(encoding='utf-8')
sys.stderr.reconfigure(encoding='utf-8')
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from schema import (
    MASTER_COLUMNS, load_rates, detect_company, detect_month, safe_float
)
from parsers.kscp_parser import KscpParser
from parsers.kctr_parser import KctrParser
from parsers.ksccz_parser import KscczParser
from parsers.ksce_parser import KsceParser
from parsers.ksci_parser import KsciParser
from parsers.ksc_parser import KscParser

PARSER_MAP = {
    "KSCP":  KscpParser,
    "KCTR":  KctrParser,
    "KSCCZ": KscczParser,
    "KSCE":  KsceParser,
    "KSCI":  KsciParser,
    "KSC":   KscParser,
}

def find_excel_files(input_dir: str, year: str = None) -> list:
    """재귀 탐색: 루트/2026/01/KCTR/*.xlsx 구조를 자동 탐지.
    단일 폴더에 파일이 있어도 동작."""
    files = []
    for root, dirs, filenames in os.walk(input_dir):
        for fn in filenames:
            if fn.startswith("~$") or "consolidated" in fn.lower():
                continue
            if fn.lower().endswith((".xlsx", ".xlsm")):
                if detect_company(fn):
                    files.append(os.path.join(root, fn))

    if year:
        year_filtered = [f for f in files if os.sep + year + os.sep in f
                         or "/" + year + "/" in f]
        if year_filtered:
            files = year_filtered

    return sorted(set(files))

def process_file(filepath: str, rates: dict) -> list:
    filename = os.path.basename(filepath)
    company = detect_company(filename)
    if not company:
        print(f"  ⚠ 법인코드 미식별: {filename}")
        return []

    parser_cls = PARSER_MAP.get(company)
    if not parser_cls:
        print(f"  ⚠ 파서 미등록: {company}")
        return []

    print(f"  📄 {filename} → {company} 파서 실행...")
    try:
        parser = parser_cls(filepath, rates)
        rows = parser.extract()
        print(f"     ✅ {len(rows)}건 추출 완료")
        return rows
    except Exception as e:
        print(f"     ❌ 파싱 실패: {e}")
        return []

def write_consolidated(all_rows: list, output_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "통합_DB"

    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="2F5496")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    for col_idx, col_name in enumerate(MASTER_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    num_fmt_local = '#,##0.00'
    num_fmt_krw = '#,##0'
    num_fmt_rate = '#,##0.00'

    for row_idx, account_row in enumerate(all_rows, 2):
        data = account_row.to_list()
        for col_idx, val in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            cell.font = Font(name="Arial", size=10)
            if col_idx == 8:
                cell.number_format = num_fmt_local
            elif col_idx == 9:
                cell.number_format = num_fmt_rate
            elif col_idx == 10:
                cell.number_format = num_fmt_krw

    col_widths = [12, 10, 6, 6, 10, 16, 8, 18, 10, 18]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    actual_rows = [r for r in all_rows if r.구분 == "실적"]
    plan_rows = [r for r in all_rows if r.구분 == "계획"]

    _create_summary(wb, actual_rows)
    _create_executive_report(wb, actual_rows)
    if plan_rows:
        _create_plan_summary(wb, plan_rows)

    wb.save(output_path)
    print(f"\n📊 통합 DB 저장: {output_path}")
    print(f"   총 {len(all_rows)}건 (실적 {len(actual_rows)}건 / 계획 {len(plan_rows)}건), {len(set(r.법인코드 for r in all_rows))}개 법인")
    sheets = "통합_DB, 법인별_요약, 경영보고서"
    if plan_rows:
        sheets += ", 사업계획_요약"
    print(f"   포함 시트: {sheets}")

def _create_summary(wb, all_rows):
    ws = wb.create_sheet("법인별_요약")

    companies = sorted(set(r.법인코드 for r in all_rows))
    periods = sorted(set(r.귀속연월 for r in all_rows))

    key_accounts = ["매출액", "매출원가", "매출총이익", "판관비계", "영업이익"]

    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill("solid", fgColor="2F5496")
    sub_fill = PatternFill("solid", fgColor="D6E4F0")

    headers = ["귀속연월", "계정과목"] + companies + ["합계"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    row_idx = 2
    for period in periods:
        for account in key_accounts:
            ws.cell(row=row_idx, column=1, value=period)
            ws.cell(row=row_idx, column=2, value=account)

            total = 0
            for comp_idx, comp in enumerate(companies, 3):
                val = sum(
                    r.KRW금액 for r in all_rows
                    if r.귀속연월 == period and r.법인코드 == comp and r.계정과목 == account
                )
                ws.cell(row=row_idx, column=comp_idx, value=val).number_format = '#,##0'
                total += val

            ws.cell(row=row_idx, column=len(companies) + 3, value=total).number_format = '#,##0'
            row_idx += 1

    for i, w in enumerate([12, 14] + [16] * (len(companies) + 1), 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    ws.freeze_panes = "C2"
    return ws

def _create_executive_report(wb, all_rows):
    ws = wb.create_sheet("경영보고서")

    companies = sorted(set(r.법인코드 for r in all_rows))
    periods = sorted(set(r.귀속연월 for r in all_rows))
    latest_period = periods[-1] if periods else "N/A"

    hdr_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    hdr_fill = PatternFill("solid", fgColor="2F5496")
    sub_hdr_fill = PatternFill("solid", fgColor="D6E4F0")
    sub_hdr_font = Font(name="Arial", bold=True, size=10)
    profit_fill = PatternFill("solid", fgColor="E2EFDA")
    loss_font = Font(name="Arial", size=10, color="FF0000")
    normal_font = Font(name="Arial", size=10)
    pct_fmt = '0.0%'
    num_fmt = '#,##0'
    thin_b = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    ws.cell(row=1, column=1, value=f"경신그룹 해외법인 경영실적 종합").font = Font(name="Arial", bold=True, size=14)
    ws.cell(row=2, column=1, value=f"최신 귀속월: {latest_period} | 누적 기간: {periods[0]} ~ {latest_period} | 단위: 백만원(KRW)").font = Font(name="Arial", size=10, color="808080")

    row = 4
    pl_accounts = [
        ("매출", ["매출액", "상품매출", "제품매출", "전선매출", "전장매출", "기타매출"]),
        ("매출원가", ["매출원가", "제품매출원가", "상품매출원가"]),
        ("이익", ["매출총이익"]),
        ("판관비", ["판관비계"]),
        ("이익", ["영업이익"]),
    ]
    mc_accounts = [
        ("재료비", ["재료비계"]),
        ("노무비", ["노무비계"]),
        ("경비", ["제조경비계", "기계경비"]),
    ]

    headers = ["구분", "계정과목"] + [f"{c}\n(당월)" for c in companies] + [f"{c}\n(누계)" for c in companies] + ["합계\n(누계)"]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=ci, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_b

    row += 1

    def get_val(corp, acct, period=None):
        if period:
            return sum(r.KRW금액 for r in all_rows if r.법인코드 == corp and r.계정과목 == acct and r.귀속연월 == period)
        return sum(r.KRW금액 for r in all_rows if r.법인코드 == corp and r.계정과목 == acct)

    def write_account_row(ws, row, section, acct, companies, latest, is_key=False):
        ws.cell(row=row, column=1, value=section).border = thin_b
        cell_acct = ws.cell(row=row, column=2, value=acct)
        cell_acct.border = thin_b
        if is_key:
            cell_acct.font = sub_hdr_font

        total_cum = 0
        for ci, comp in enumerate(companies):
            monthly = get_val(comp, acct, latest) / 1_000_000
            cum = get_val(comp, acct) / 1_000_000

            m_cell = ws.cell(row=row, column=3 + ci, value=round(monthly, 0))
            m_cell.number_format = num_fmt
            m_cell.border = thin_b
            m_cell.font = loss_font if monthly < 0 else normal_font

            c_cell = ws.cell(row=row, column=3 + len(companies) + ci, value=round(cum, 0))
            c_cell.number_format = num_fmt
            c_cell.border = thin_b
            c_cell.font = loss_font if cum < 0 else normal_font

            total_cum += cum

        t_cell = ws.cell(row=row, column=3 + 2 * len(companies), value=round(total_cum, 0))
        t_cell.number_format = num_fmt
        t_cell.border = thin_b
        t_cell.font = Font(name="Arial", bold=True, size=10, color="FF0000" if total_cum < 0 else "000000")

        if acct in ("매출총이익", "영업이익"):
            for c in range(1, 3 + 2 * len(companies) + 1):
                ws.cell(row=row, column=c).fill = profit_fill

    ws.cell(row=row, column=1, value="【 PL 】").font = sub_hdr_font
    ws.cell(row=row, column=1).fill = sub_hdr_fill
    for ci in range(2, 3 + 2 * len(companies) + 1):
        ws.cell(row=row, column=ci).fill = sub_hdr_fill
        ws.cell(row=row, column=ci).border = thin_b
    row += 1

    for section, accts in pl_accounts:
        for acct in accts:
            has_data = any(get_val(c, acct) != 0 for c in companies)
            if has_data:
                is_key = acct in ("매출액", "매출원가", "매출총이익", "판관비계", "영업이익")
                write_account_row(ws, row, section, acct, companies, latest_period, is_key)
                row += 1

    row += 1
    ws.cell(row=row, column=1, value="【 제조원가 】").font = sub_hdr_font
    ws.cell(row=row, column=1).fill = sub_hdr_fill
    for ci in range(2, 3 + 2 * len(companies) + 1):
        ws.cell(row=row, column=ci).fill = sub_hdr_fill
        ws.cell(row=row, column=ci).border = thin_b
    row += 1

    for section, accts in mc_accounts:
        for acct in accts:
            has_data = any(get_val(c, acct) != 0 for c in companies)
            if has_data:
                write_account_row(ws, row, section, acct, companies, latest_period)
                row += 1

    row += 1
    ws.cell(row=row, column=1, value="【 이익률 (%) 】").font = sub_hdr_font
    ws.cell(row=row, column=1).fill = sub_hdr_fill
    for ci in range(2, 3 + 2 * len(companies) + 1):
        ws.cell(row=row, column=ci).fill = sub_hdr_fill
    row += 1

    for label in ["매출총이익률", "영업이익률"]:
        ws.cell(row=row, column=1, value="이익률").border = thin_b
        ws.cell(row=row, column=2, value=label).border = thin_b
        ws.cell(row=row, column=2).font = sub_hdr_font

        acct = "매출총이익" if "총" in label else "영업이익"
        for ci, comp in enumerate(companies):
            rev_m = get_val(comp, "매출액", latest_period)
            profit_m = get_val(comp, acct, latest_period)
            rev_c = get_val(comp, "매출액")
            profit_c = get_val(comp, acct)

            m_cell = ws.cell(row=row, column=3 + ci, value=profit_m / rev_m if rev_m else 0)
            m_cell.number_format = pct_fmt
            m_cell.border = thin_b
            c_cell = ws.cell(row=row, column=3 + len(companies) + ci, value=profit_c / rev_c if rev_c else 0)
            c_cell.number_format = pct_fmt
            c_cell.border = thin_b

        row += 1

    col_widths = [10, 16] + [14] * (2 * len(companies) + 1)
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=4, column=i).column_letter].width = w

    ws.freeze_panes = "C5"
    ws.sheet_properties.tabColor = "2F5496"

def _create_plan_summary(wb, plan_rows):
    ws = wb.create_sheet("사업계획_요약")

    companies = sorted(set(r.법인코드 for r in plan_rows))
    periods = sorted(set(r.귀속연월 for r in plan_rows))
    current_ym = datetime.now().strftime("%Y-%m")
    latest_period = current_ym

    key_accounts = ["매출액", "매출원가", "매출총이익", "판관비계", "영업이익"]

    hdr_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    hdr_fill = PatternFill("solid", fgColor="375623")
    sub_hdr_font = Font(name="Arial", bold=True, size=10)
    profit_fill = PatternFill("solid", fgColor="E2EFDA")
    normal_font = Font(name="Arial", size=10)
    loss_font = Font(name="Arial", size=10, color="FF0000")
    num_fmt = '#,##0'
    thin_b = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    ws.cell(row=1, column=1, value="경신그룹 해외법인 사업계획 종합").font = Font(name="Arial", bold=True, size=14)
    ws.cell(row=2, column=1, value=f"당월: {latest_period} | 누적: {periods[0]} ~ {latest_period} | 전체: 연간계획 합계 | 단위: 백만원(KRW)").font = Font(name="Arial", size=10, color="808080")

    row = 4
    # 헤더: 구분, 계정과목, [법인별 당월계획], [법인별 누적계획], [법인별 전체계획], 합계(전체)
    headers = (
        ["구분", "계정과목"]
        + [f"{c}\n(당월)" for c in companies]
        + [f"{c}\n(누계)" for c in companies]
        + [f"{c}\n(전체)" for c in companies]
        + ["합계\n(전체)"]
    )
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=ci, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_b
    row += 1

    def get_plan(corp, acct, period=None, cumulative=False):
        if period:
            return sum(r.KRW금액 for r in plan_rows if r.법인코드 == corp and r.계정과목 == acct and r.귀속연월 == period)
        elif cumulative:
            return sum(r.KRW금액 for r in plan_rows if r.법인코드 == corp and r.계정과목 == acct and r.귀속연월 <= latest_period)
        else:
            return sum(r.KRW금액 for r in plan_rows if r.법인코드 == corp and r.계정과목 == acct)

    n = len(companies)
    for acct in key_accounts:
        has_data = any(get_plan(c, acct) != 0 for c in companies)
        if not has_data:
            continue

        is_key = acct in ("매출액", "매출원가", "매출총이익", "판관비계", "영업이익")
        cell_acct = ws.cell(row=row, column=2, value=acct)
        cell_acct.border = thin_b
        if is_key:
            cell_acct.font = sub_hdr_font
        ws.cell(row=row, column=1).border = thin_b

        total_full = 0
        for ci, comp in enumerate(companies):
            monthly = get_plan(comp, acct, period=latest_period) / 1_000_000
            cum = get_plan(comp, acct, cumulative=True) / 1_000_000
            full = get_plan(comp, acct) / 1_000_000

            for col_offset, val in ((0, monthly), (n, cum), (2 * n, full)):
                cell = ws.cell(row=row, column=3 + ci + col_offset, value=round(val, 0))
                cell.number_format = num_fmt
                cell.border = thin_b
                cell.font = loss_font if val < 0 else normal_font

            total_full += full

        t_cell = ws.cell(row=row, column=3 + 3 * n, value=round(total_full, 0))
        t_cell.number_format = num_fmt
        t_cell.border = thin_b
        t_cell.font = Font(name="Arial", bold=True, size=10, color="FF0000" if total_full < 0 else "000000")

        if acct in ("매출총이익", "영업이익"):
            for c in range(1, 3 + 3 * n + 1):
                ws.cell(row=row, column=c).fill = profit_fill

        row += 1

    col_widths = [10, 16] + [14] * (3 * n + 1)
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=4, column=i).column_letter].width = w

    ws.freeze_panes = "C5"
    ws.sheet_properties.tabColor = "375623"
    return ws


def load_settings(config_dir: str = None) -> dict:
    if config_dir is None:
        config_dir = os.path.join(os.path.dirname(__file__), "config")
    path = os.path.join(config_dir, "settings.json")
    if os.path.exists(path):
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}

def save_settings(settings: dict, config_dir: str = None):
    if config_dir is None:
        config_dir = os.path.join(os.path.dirname(__file__), "config")
    os.makedirs(config_dir, exist_ok=True)
    path = os.path.join(config_dir, "settings.json")
    with open(path, "w", encoding="utf-8") as f:
        json.dump(settings, f, ensure_ascii=False, indent=2)

def _get_base_dir() -> str:
    """Returns the install directory where config/ lives (may be read-only).
    In a PyInstaller --onedir bundle, config/ is next to engine.exe."""
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))

def _get_user_config_dir() -> str:
    # Returns a user-writable config dir: %APPDATA%/KSC Refiner/
    appdata = os.environ.get("APPDATA") or os.path.expanduser("~")
    return os.path.join(appdata, "KSC Refiner")


def main(input_dir: str = None, output_dir: str = None, year: str = "2026"):
    install_config = os.path.join(_get_base_dir(), "config")
    user_config = _get_user_config_dir()
    os.makedirs(user_config, exist_ok=True)

    # Copy rates.json to user config on first run so user edits persist
    user_rates = os.path.join(user_config, "rates.json")
    install_rates = os.path.join(install_config, "rates.json")
    if not os.path.exists(user_rates) and os.path.exists(install_rates):
        import shutil
        shutil.copy2(install_rates, user_rates)

    # settings.json is always read/written in user config (writable)
    settings = load_settings(user_config)

    if input_dir is None:
        input_dir = settings.get("root_path", os.path.join(os.path.dirname(__file__), "..", "input"))
    if output_dir is None:
        saved = settings.get("output_path", "")
        # 저장된 경로가 같은 루트 하위이거나 비어 있으면 input 옆 output 폴더 사용
        if saved and os.path.splitdrive(saved)[0].lower() == os.path.splitdrive(input_dir)[0].lower():
            try:
                os.makedirs(saved, exist_ok=True)
                output_dir = saved
            except (PermissionError, OSError):
                output_dir = os.path.join(input_dir, "output")
        else:
            output_dir = os.path.join(input_dir, "output")

    os.makedirs(output_dir, exist_ok=True)

    settings["root_path"] = input_dir
    settings["output_path"] = output_dir
    settings["last_year"] = year
    save_settings(settings, user_config)

    print("=" * 60)
    print("🚀 KSC Settlement Refiner Engine v1.2")
    print(f"   루트 경로: {input_dir}")
    print(f"   기준 연도: {year}")
    print(f"   처리 범위: {year}년 전체 누적")
    print("=" * 60)

    # Load rates from user config (writable), fall back to install config
    rates_config = user_config if os.path.exists(user_rates) else install_config
    rates = load_rates(year, rates_config)
    print(f"\n📌 적용 환율: {rates}")

    files = find_excel_files(input_dir, year)
    if not files:
        print("\n❌ 처리할 엑셀 파일이 없소.")
        return None

    print(f"\n📂 탐지된 파일: {len(files)}개")
    for f in files:
        print(f"   {os.path.relpath(f, input_dir)}")

    all_rows = []
    for filepath in sorted(files):
        rows = process_file(filepath, rates)
        all_rows.extend(rows)

    if not all_rows:
        print("\n❌ 추출된 데이터가 없소.")
        return None

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = os.path.join(output_dir, f"consolidated_db_{timestamp}.xlsx")
    write_consolidated(all_rows, output_path)
    return output_path

if __name__ == "__main__":
    input_dir = sys.argv[1] if len(sys.argv) > 1 else None
    year = sys.argv[2] if len(sys.argv) > 2 else "2026"
    main(input_dir=input_dir, year=year)
