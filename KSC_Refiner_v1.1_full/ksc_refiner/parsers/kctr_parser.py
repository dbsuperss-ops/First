import openpyxl
from typing import List
from parsers import BaseParser
from schema import AccountRow, safe_float, CATEGORY_PL, CATEGORY_MC

KCTR_PL_MAP = {
    5:  (CATEGORY_PL, "매출", "매출액"),
    6:  (CATEGORY_PL, "매출", "상품매출"),
    16: (CATEGORY_PL, "매출", "기타매출"),
    24: (CATEGORY_PL, "매출원가", "매출원가"),
    26: (CATEGORY_PL, "매출원가", "상품매출원가"),
    28: (CATEGORY_PL, "매출원가", "기초상품재고"),
    29: (CATEGORY_PL, "매출원가", "당기상품매입원가"),
    32: (CATEGORY_PL, "매출원가", "기말상품재고"),
    33: (CATEGORY_PL, "매출원가", "기타매출원가"),
    38: (CATEGORY_PL, "이익", "매출총이익"),
    40: (CATEGORY_PL, "판관비", "판매비"),
    48: (CATEGORY_PL, "판관비", "관리비"),
    66: (CATEGORY_PL, "이익", "영업이익"),
}

KCTR_MC_MAP = {
    62: (CATEGORY_MC, "재료비", "재료비계"),
    63: (CATEGORY_MC, "노무비", "노무비계"),
    64: (CATEGORY_MC, "경비", "제조경비계"),
    65: (CATEGORY_MC, "경비", "감가상각비"),
}

MONTH_COL_START = 6  # F column = 1월

# 사업계획 파일: Summary 시트 (col 5=1월 ~ col 16=12월)
KCTR_PLAN_MAP = {
    5:  (CATEGORY_PL, "매출",    "매출액"),
    20: (CATEGORY_PL, "매출원가", "매출원가"),
    32: (CATEGORY_PL, "이익",    "매출총이익"),
    34: (CATEGORY_PL, "판관비",  "판매비"),
    43: (CATEGORY_PL, "판관비",  "관리비"),
    70: (CATEGORY_PL, "이익",    "영업이익"),
}
KCTR_PLAN_MONTH_COL_START = 5  # E column = 1월

class KctrParser(BaseParser):
    def extract(self) -> List[AccountRow]:
        wb = openpyxl.load_workbook(self.filepath, read_only=True, data_only=True)
        rows = []

        if "PL" not in wb.sheetnames:
            if "Summary" in wb.sheetnames:
                rows = self._extract_plan(wb)
            else:
                print(f"  [KCTR] 지원 시트 없음 (PL, Summary 모두 없음) — 건너뜀")
            wb.close()
            return rows

        ws_pl = wb["PL"]
        year = self._detect_year(ws_pl)
        pl_map = dict(KCTR_PL_MAP)

        mc_rows_exist = False
        for r in KCTR_MC_MAP:
            v = ws_pl.cell(row=r, column=2).value
            if v and any(kw in str(v) for kw in ["재료", "노무", "경비", "감가"]):
                mc_rows_exist = True
                break

        if mc_rows_exist:
            pl_map.update(KCTR_MC_MAP)

        for month_offset in range(12):
            col = MONTH_COL_START + month_offset
            month_num = month_offset + 1

            test_val = ws_pl.cell(row=5, column=col).value
            if test_val is None or safe_float(test_val) == 0:
                continue

            ym = f"{year}-{month_num:02d}"
            for row_num, (cat, sub, account) in pl_map.items():
                val = safe_float(ws_pl.cell(row=row_num, column=col).value)
                rows.append(self.make_row(ym, "KCTR", cat, sub, account, "TRY", val))

        if not mc_rows_exist:
            rows.extend(self._extract_from_report(wb, year))

        wb.close()
        return rows

    def _extract_from_report(self, wb, year) -> List[AccountRow]:
        if "★PL(Report)" not in wb.sheetnames:
            return []
        ws = wb["★PL(Report)"]
        month_cell = ws.cell(row=4, column=3).value
        if month_cell is None:
            return []
        month_num = int(safe_float(month_cell))
        if month_num == 0:
            return []
        ym = f"{year}-{month_num:02d}"

        report_rows = []
        for r in range(50, min(ws.max_row + 1, 100)):
            label = ws.cell(row=r, column=2).value
            if not label:
                continue
            label_str = str(label).strip()
            actual_val = safe_float(ws.cell(row=r, column=5).value)  # E=누적실적
            if "재료비" in label_str:
                report_rows.append(self.make_row(ym, "KCTR", CATEGORY_MC, "재료비", "재료비계", "TRY", actual_val))
            elif "노무비" in label_str:
                report_rows.append(self.make_row(ym, "KCTR", CATEGORY_MC, "노무비", "노무비계", "TRY", actual_val))
            elif "제조경비" in label_str or "경비" in label_str:
                report_rows.append(self.make_row(ym, "KCTR", CATEGORY_MC, "경비", "제조경비계", "TRY", actual_val))
        return report_rows

    def _extract_plan(self, wb) -> List[AccountRow]:
        """사업계획 파일: Summary 시트에서 월별 데이터 추출"""
        ws = wb["Summary"]
        year = self._detect_year_plan(ws)
        rows = []
        for month_offset in range(12):
            col = KCTR_PLAN_MONTH_COL_START + month_offset
            month_num = month_offset + 1
            test_val = ws.cell(row=5, column=col).value
            if test_val is None or safe_float(test_val) == 0:
                continue
            ym = f"{year}-{month_num:02d}"
            for row_num, (cat, sub, account) in KCTR_PLAN_MAP.items():
                val = safe_float(ws.cell(row=row_num, column=col).value)
                rows.append(self.make_row(ym, "KCTR", cat, sub, account, "TRY", val))
        return rows

    def _detect_year_plan(self, ws) -> str:
        import re
        for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
            for cell in row:
                if cell:
                    m = re.search(r'(\d{4})', str(cell))
                    if m and m.group(1) in ("2024", "2025", "2026", "2027"):
                        return m.group(1)
        return "2026"

    def _detect_year(self, ws) -> str:
        for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
            for cell in row:
                if cell and "2026" in str(cell):
                    return "2026"
                if cell and "2025" in str(cell):
                    return "2025"
        return "2026"
