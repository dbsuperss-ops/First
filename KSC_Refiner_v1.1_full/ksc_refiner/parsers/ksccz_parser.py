import openpyxl
import re
from typing import List
from parsers import BaseParser
from schema import AccountRow, safe_float, CATEGORY_PL, CATEGORY_MC

KSCCZ_PL_MAP = {
    3:  (CATEGORY_PL, "매출", "매출액"),
    4:  (CATEGORY_PL, "매출", "제품매출"),
    5:  (CATEGORY_PL, "매출", "상품매출"),
    6:  (CATEGORY_PL, "매출", "기타매출"),
    10: (CATEGORY_PL, "매출원가", "매출원가"),
    11: (CATEGORY_PL, "매출원가", "제품매출원가"),
    12: (CATEGORY_PL, "매출원가", "상품매출원가"),
    16: (CATEGORY_PL, "이익", "매출총이익"),
    18: (CATEGORY_PL, "판관비", "판관비계"),
    19: (CATEGORY_PL, "판관비", "급여"),
    20: (CATEGORY_PL, "판관비", "복리후생비"),
    21: (CATEGORY_PL, "판관비", "운반비"),
    22: (CATEGORY_PL, "판관비", "여비교통비"),
    23: (CATEGORY_PL, "판관비", "지급수수료"),
    24: (CATEGORY_PL, "판관비", "통신비"),
    25: (CATEGORY_PL, "판관비", "접대비"),
    26: (CATEGORY_PL, "판관비", "감가상각비"),
}

KSCCZ_MC_MAP = {
    3:  (CATEGORY_MC, "재료비", "재료비계"),
    4:  (CATEGORY_MC, "재료비", "기초재고"),
    5:  (CATEGORY_MC, "재료비", "당기매입액"),
    7:  (CATEGORY_MC, "재료비", "기말재고액"),
    8:  (CATEGORY_MC, "노무비", "노무비계"),
    10: (CATEGORY_MC, "경비", "제조경비계"),
    11: (CATEGORY_MC, "경비", "복리후생비"),
    15: (CATEGORY_MC, "경비", "수도광열비"),
    17: (CATEGORY_MC, "경비", "감가상각비"),
    19: (CATEGORY_MC, "경비", "수선비"),
    20: (CATEGORY_MC, "경비", "보험료"),
}

ACTUAL_COL = 7  # ★손익계산서: G column = 실적

# 손익계산서 (월별) 시트: 사업부 섹션 시작 행 및 월 열 시작
MONTHLY_SECTION_START = 161   # 전장사업부 섹션 시작 행 (161행 이하)
MONTHLY_COL_START = 5         # E열 = 1월

# 손익계산서 (월별) 시트용 레이블 매핑 (exact match → contains 순서로 탐색)
MONTHLY_LABEL_MAP = {
    "매출액":      (CATEGORY_PL, "매출",    "매출액"),
    "제품매출":    (CATEGORY_PL, "매출",    "제품매출"),
    "상품매출":    (CATEGORY_PL, "매출",    "상품매출"),
    "기타매출":    (CATEGORY_PL, "매출",    "기타매출"),
    "매출원가":    (CATEGORY_PL, "매출원가", "매출원가"),
    "제품매출원가": (CATEGORY_PL, "매출원가", "제품매출원가"),
    "상품매출원가": (CATEGORY_PL, "매출원가", "상품매출원가"),
    "매출총이익":  (CATEGORY_PL, "이익",    "매출총이익"),
    "판관비계":    (CATEGORY_PL, "판관비",  "판관비계"),
    "판매관리비":  (CATEGORY_PL, "판관비",  "판관비계"),  # alias: sheet uses "IV.판매관리비"
    "영업이익":    (CATEGORY_PL, "이익",    "영업이익"),
}

# 사업계획 파일: 손익계획 시트 — 전장사업부 섹션은 180행 이하, 레이블 기반 추출
KSCCZ_PLAN_SECTION_START = 180
KSCCZ_PLAN_LABEL_MAP = {
    "매출액":      (CATEGORY_PL, "매출",    "매출액"),
    "매출원가":    (CATEGORY_PL, "매출원가", "매출원가"),
    "매출총이익":  (CATEGORY_PL, "이익",    "매출총이익"),
    "판관비계":    (CATEGORY_PL, "판관비",  "판관비계"),
    "판매관리비":  (CATEGORY_PL, "판관비",  "판관비계"),
    "영업이익":    (CATEGORY_PL, "이익",    "영업이익"),
}
# 사업계획 파일: 제조원가계획 시트 (col 5=1월 ~ col 16=12월)
KSCCZ_PLAN_MC_MAP = {
    4:  (CATEGORY_MC, "재료비", "재료비계"),
    11: (CATEGORY_MC, "노무비", "노무비계"),
    16: (CATEGORY_MC, "경비",  "제조경비계"),
}
KSCCZ_PLAN_MONTH_COL_START = 5  # E column = 1월


class KscczParser(BaseParser):
    def extract(self) -> List[AccountRow]:
        wb = openpyxl.load_workbook(self.filepath, read_only=True, data_only=True)
        rows = []

        # 우선순위 1: 손익계산서 (월별) 시트 → 사업부 섹션(161행~)
        monthly_sheet = next(
            (s for s in wb.sheetnames if "손익계산서" in s and "월별" in s), None
        )
        if monthly_sheet:
            year = self._detect_year_from_wb(wb)
            rows.extend(self._extract_monthly_sheet(wb[monthly_sheet], year))

        # 우선순위 2: ★손익계산서 (단월 실적 파일)
        elif "★손익계산서" in wb.sheetnames:
            ws_pl = wb["★손익계산서"]
            month_num = self._detect_month(ws_pl)
            year = self._detect_year(ws_pl)
            ym = f"{year}-{month_num:02d}"

            for row_num, (cat, sub, account) in KSCCZ_PL_MAP.items():
                val = safe_float(ws_pl.cell(row=row_num, column=ACTUAL_COL).value)
                rows.append(self.make_row(ym, "KSCCZ", cat, sub, account, "RMB", val))

            op_profit = (
                safe_float(ws_pl.cell(row=16, column=ACTUAL_COL).value) -
                safe_float(ws_pl.cell(row=18, column=ACTUAL_COL).value)
            )
            rows.append(self.make_row(ym, "KSCCZ", CATEGORY_PL, "이익", "영업이익", "RMB", op_profit))

            if "★제조원가명세서" in wb.sheetnames:
                ws_mc = wb["★제조원가명세서"]
                for row_num, (cat, sub, account) in KSCCZ_MC_MAP.items():
                    val = safe_float(ws_mc.cell(row=row_num, column=ACTUAL_COL).value)
                    rows.append(self.make_row(ym, "KSCCZ", cat, sub, account, "RMB", val))

        # 우선순위 3: 사업계획 파일
        elif "손익계획" in wb.sheetnames:
            rows = self._extract_plan(wb)

        else:
            print(f"  [KSCCZ] 지원 시트 없음 (손익계산서(월별), ★손익계산서, 손익계획 모두 없음) — 건너뜀")

        wb.close()
        return rows

    def _extract_monthly_sheet(self, ws, year: str) -> List[AccountRow]:
        # 손익계산서 (월별) 시트: 전장사업부 데이터는 161행부터 시작 (이전 행은 타 사업부 자료)
        label_rows: dict = {}

        start_row = MONTHLY_SECTION_START
        scan_end = min(ws.max_row + 1, start_row + 400)

        for r in range(start_row, scan_end):
            for c in (1, 2, 3, 4, 5):
                raw = ws.cell(row=r, column=c).value
                if raw is None:
                    continue
                stripped = str(raw).replace(" ", "").replace("\n", "").strip()
                import re
                clean_name = re.sub(r'^[\dIVXivx\.\-\s]+', '', stripped)
                for label in MONTHLY_LABEL_MAP:
                    if label not in label_rows and (
                        clean_name == label or stripped == label or
                        (len(label) >= 3 and stripped.endswith(label))
                    ):
                        label_rows[label] = r
                        break

        if "매출액" not in label_rows:
            # 진단: 발견된 레이블과 시트 내 A~E 컬럼 처음 20개 텍스트 출력
            found = list(label_rows.keys()) or []
            samples = []
            for r in range(start_row, min(start_row + 30, ws.max_row + 1)):
                for c in (1, 2, 3):
                    v = ws.cell(row=r, column=c).value
                    if v:
                        samples.append(f"  row{r}c{c}={repr(str(v)[:30])}")
                        if len(samples) >= 15:
                            break
                if len(samples) >= 15:
                    break
            print(f"  [KSCCZ] 손익계산서(월별) (탐색시작: {start_row}행~, 종료: {scan_end}행) 매출액 미발견")
            if found:
                print(f"  [KSCCZ] 발견된 레이블: {found}")
            if samples:
                print(f"  [KSCCZ] 탐색 범위 샘플:")
                for s in samples:
                    print(f"  {s}")
            return []

        sales_row = label_rows["매출액"]
        rows = []

        for month_offset in range(12):
            col = MONTHLY_COL_START + month_offset
            month_num = month_offset + 1
            if safe_float(ws.cell(row=sales_row, column=col).value) == 0:
                continue
            ym = f"{year}-{month_num:02d}"
            for label, (cat, sub, account) in MONTHLY_LABEL_MAP.items():
                if label not in label_rows:
                    continue
                val = safe_float(ws.cell(row=label_rows[label], column=col).value)
                rows.append(self.make_row(ym, "KSCCZ", cat, sub, account, "RMB", val))

        return rows

    def _extract_plan(self, wb) -> List[AccountRow]:
        """사업계획 파일: 손익계획 + 제조원가계획 시트에서 월별 데이터 추출 (레이블 기반)"""
        ws_pl = wb["손익계획"]
        year = self._detect_year(ws_pl)
        rows = []

        # 180행 이하에서 레이블 기반으로 행 번호 탐색
        label_rows: dict = {}
        start_row = KSCCZ_PLAN_SECTION_START
        scan_end = min(ws_pl.max_row + 1, start_row + 400)

        for r in range(start_row, scan_end):
            for c in (1, 2, 3, 4):
                raw = ws_pl.cell(row=r, column=c).value
                if raw is None:
                    continue
                stripped = str(raw).replace(" ", "").replace("\n", "").strip()
                import re as _re
                clean_name = _re.sub(r'^[\dIVXivx\.\-\s]+', '', stripped)
                for label in KSCCZ_PLAN_LABEL_MAP:
                    if label not in label_rows and (
                        clean_name == label or stripped == label or
                        (len(label) >= 3 and stripped.endswith(label))
                    ):
                        label_rows[label] = r
                        break

        if "매출액" not in label_rows:
            print(f"  [KSCCZ] 손익계획 시트 (탐색시작: {start_row}행~) 매출액 미발견 — 건너뜀")
            return []

        for month_offset in range(12):
            col = KSCCZ_PLAN_MONTH_COL_START + month_offset
            month_num = month_offset + 1
            test_val = ws_pl.cell(row=label_rows["매출액"], column=col).value
            if test_val is None or safe_float(test_val) == 0:
                continue
            ym = f"{year}-{month_num:02d}"

            for label, (cat, sub, account) in KSCCZ_PLAN_LABEL_MAP.items():
                if label not in label_rows:
                    continue
                val = safe_float(ws_pl.cell(row=label_rows[label], column=col).value)
                rows.append(self.make_row(ym, "KSCCZ", cat, sub, account, "RMB", val, 구분="계획"))

            if "제조원가계획" in wb.sheetnames:
                ws_mc = wb["제조원가계획"]
                for row_num, (cat, sub, account) in KSCCZ_PLAN_MC_MAP.items():
                    val = safe_float(ws_mc.cell(row=row_num, column=col).value)
                    rows.append(self.make_row(ym, "KSCCZ", cat, sub, account, "RMB", val, 구분="계획"))

        return rows

    def _detect_year_from_wb(self, wb) -> str:
        """워크북 첫 번째 시트에서 연도 탐지"""
        for sname in wb.sheetnames:
            ws = wb[sname]
            year = self._detect_year(ws)
            if year != "2026":
                return year
            break
        return "2026"

    def _detect_month(self, ws) -> int:
        val = ws.cell(row=1, column=2).value
        if val and isinstance(val, (int, float)):
            return int(val)
        for row in ws.iter_rows(min_row=1, max_row=3, values_only=True):
            for cell in row:
                if cell:
                    m = re.search(r'(\d{1,2})\s*월', str(cell))
                    if m:
                        return int(m.group(1))
        return 1

    def _detect_year(self, ws) -> str:
        for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
            for cell in row:
                if cell:
                    m = re.search(r'(\d{4})', str(cell))
                    if m and m.group(1) in ("2024", "2025", "2026", "2027"):
                        return m.group(1)
        return "2026"
