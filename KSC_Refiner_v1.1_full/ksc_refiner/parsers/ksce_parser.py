import openpyxl
import re
from typing import List
from parsers import BaseParser
from schema import AccountRow, safe_float, CATEGORY_PL, CATEGORY_MC

# PL 시트: col B 레이블 → (대분류, 중분류, 계정과목)
PL_LABEL_MAP = {
    "Ⅰ. 매출액":     (CATEGORY_PL, "매출",   "매출액"),
    "Ⅱ. 매출원가":   (CATEGORY_PL, "매출원가", "매출원가"),
    "Ⅲ. 매출총이익": (CATEGORY_PL, "이익",   "매출총이익"),
    "Ⅳ.판매관리비":  (CATEGORY_PL, "판관비",  "판관비계"),
    "Ⅴ.영업이익":    (CATEGORY_PL, "이익",   "영업이익"),
}

# MC 시트: col B 레이블 → (대분류, 중분류, 계정과목)
MC_LABEL_MAP = {
    "Ⅰ. 재료비": (CATEGORY_MC, "재료비", "재료비계"),
    "Ⅱ. 노무비": (CATEGORY_MC, "노무비", "노무비계"),
    "Ⅲ. 경비":   (CATEGORY_MC, "경비",  "제조경비계"),
}

# PL 시트: 월 데이터 시작 열 (F열 = 1월)
PL_MONTH_COL_START = 6
# MC 시트: 월 데이터 시작 열 (E열 = 1월)
MC_MONTH_COL_START = 5

PLAN_MONTH_COL_START = 23  # PL2026(RS/KR): W열 = 1월 (2026 사업계획)

# IS * 시트 (KSCI 공용 템플릿): col G = 1월
IS_MONTH_COL_START = 7
KSCE_IS_MAP = {
    5:  (CATEGORY_PL, "매출",    "매출액"),
    9:  (CATEGORY_MC, "재료비",  "재료비계"),
    21: (CATEGORY_MC, "노무비",  "노무비계"),
    35: (CATEGORY_MC, "경비",    "제조경비계"),
    71: (CATEGORY_PL, "판관비",  "판관비계"),
}


class KsceParser(BaseParser):
    def extract(self) -> List[AccountRow]:
        wb = openpyxl.load_workbook(self.filepath, read_only=True, data_only=True)
        rows = []
        year = self._detect_year(wb)

        # 사업계획 파일: PL2026(RS) 또는 PL2026(KR) 시트
        plan_sheet = None
        for sname in wb.sheetnames:
            if sname.startswith("PL2026"):
                plan_sheet = sname
                break

        if plan_sheet:
            # 시트명 PL2026(RS) / PL2026(KR) 에서 연도 추출
            m_year = re.search(r"PL(\d{4})", plan_sheet)
            plan_year = m_year.group(1) if m_year else year
            rows.extend(self._extract_sheet(
                wb[plan_sheet], plan_year, PL_LABEL_MAP, PLAN_MONTH_COL_START, label_col=1, 구분="계획"
            ))
        elif "PL" in wb.sheetnames or "MC" in wb.sheetnames:
            # 실적 파일: PL / MC 시트
            if "PL" in wb.sheetnames:
                rows.extend(self._extract_sheet(
                    wb["PL"], year, PL_LABEL_MAP, PL_MONTH_COL_START, label_col=2
                ))
            if "MC" in wb.sheetnames:
                rows.extend(self._extract_sheet(
                    wb["MC"], year, MC_LABEL_MAP, MC_MONTH_COL_START, label_col=2
                ))
        else:
            # IS * 시트 (KSCI 공용 템플릿)
            is_sheets = [s for s in wb.sheetnames if s.startswith("IS ")]
            if is_sheets:
                rows.extend(self._extract_is_sheets(wb, is_sheets, year))
            else:
                print(f"  [KSCE] 지원 시트 없음 (PL2026*, PL/MC, IS* 모두 없음) — 건너뜀")

        wb.close()
        return rows

    def _extract_sheet(self, ws, year: str, label_map: dict, month_col_start: int, label_col: int = 2, 구분: str = "실적") -> List[AccountRow]:
        # label_col 열에서 레이블 위치 탐색
        label_rows = {}
        for r in range(1, ws.max_row + 1):
            raw = ws.cell(row=r, column=label_col).value
            if raw is None:
                continue
            key = str(raw).strip()
            if key in label_map:
                label_rows[key] = r

        rows = []
        for month_offset in range(12):
            col = month_col_start + month_offset
            month_num = month_offset + 1

            # 해당 월에 데이터가 있는지 확인
            has_data = any(
                safe_float(ws.cell(row=row_num, column=col).value) != 0
                for row_num in label_rows.values()
            )
            if not has_data:
                continue

            ym = f"{year}-{month_num:02d}"
            for label, (cat, sub, account) in label_map.items():
                if label not in label_rows:
                    continue
                val = safe_float(ws.cell(row=label_rows[label], column=col).value)
                rows.append(self.make_row(ym, "KSCE", cat, sub, account, "RSD", val, 구분=구분))

        return rows

    def _extract_is_sheets(self, wb, sheet_names: list, year: str) -> List[AccountRow]:
        """IS * 시트들을 합산하여 월별 데이터 추출 (KSCI 공용 템플릿)"""
        monthly = {}
        for sname in sheet_names:
            ws = wb[sname]
            for month_offset in range(12):
                col = IS_MONTH_COL_START + month_offset
                test_val = ws.cell(row=5, column=col).value
                if test_val is None:
                    continue
                for row_num in KSCE_IS_MAP:
                    key = (month_offset, row_num)
                    monthly[key] = monthly.get(key, 0.0) + safe_float(ws.cell(row=row_num, column=col).value)

        rows = []
        for month_offset in range(12):
            month_num = month_offset + 1
            ym = f"{year}-{month_num:02d}"
            sales = monthly.get((month_offset, 5), 0.0)
            if sales == 0:
                continue
            for row_num, (cat, sub, account) in KSCE_IS_MAP.items():
                val = monthly.get((month_offset, row_num), 0.0)
                rows.append(self.make_row(ym, "KSCE", cat, sub, account, "USD", val))
            gross = (monthly.get((month_offset, 5), 0.0)
                     - monthly.get((month_offset, 9), 0.0)
                     - monthly.get((month_offset, 21), 0.0)
                     - monthly.get((month_offset, 35), 0.0))
            rows.append(self.make_row(ym, "KSCE", CATEGORY_PL, "이익", "매출총이익", "USD", gross))
            op = gross - monthly.get((month_offset, 71), 0.0)
            rows.append(self.make_row(ym, "KSCE", CATEGORY_PL, "이익", "영업이익", "USD", op))
        return rows

    def _detect_year(self, wb) -> str:
        check_sheets = ["PL", "PL(Report)"]
        # 사업계획 시트도 연도 탐색에 포함
        for sname in wb.sheetnames:
            if sname.startswith("PL2026"):
                check_sheets.insert(0, sname)
                break

        for sname in check_sheets:
            if sname not in wb.sheetnames:
                continue
            ws = wb[sname]
            for r in range(1, 8):
                for c in range(1, 10):
                    val = ws.cell(row=r, column=c).value
                    if not val or not isinstance(val, str):
                        continue
                    m = re.search(r"FY\s*(\d{4})", val)
                    if m:
                        return m.group(1)
                    m = re.search(r"(\d{4})", val)
                    if m and m.group(1) in ("2024", "2025", "2026", "2027"):
                        return m.group(1)
        # 파일명에서 연도 추출 (fallback)
        m = re.search(r"(\d{4})", self.filepath)
        if m:
            return m.group(1)
        return "2026"
