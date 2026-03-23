import openpyxl
from typing import List
from parsers import BaseParser
from schema import AccountRow, safe_float, CATEGORY_PL, CATEGORY_MC

KSCP_ACCOUNT_MAP = {
    2:  (CATEGORY_PL, "매출", "매출액"),
    3:  (CATEGORY_PL, "매출원가", "매출원가"),
    4:  (CATEGORY_PL, "이익", "매출총이익"),
    5:  (CATEGORY_PL, "이익", "영업이익"),
    6:  (CATEGORY_MC, "재료비", "재료비계"),
    7:  (CATEGORY_MC, "재료비", "원재료"),
    8:  (CATEGORY_MC, "재료비", "관세"),
    9:  (CATEGORY_MC, "재료비", "통관물류"),
    10: (CATEGORY_MC, "재료비", "운반비(고객)"),
    11: (CATEGORY_MC, "재료비", "재고평가손실"),
    12: (CATEGORY_MC, "노무비", "노무비계"),
    13: (CATEGORY_MC, "노무비", "직접노무비"),
    14: (CATEGORY_MC, "노무비", "간접노무비"),
    15: (CATEGORY_MC, "경비", "기계경비"),
    16: (CATEGORY_PL, "판관비", "판관비계"),
    17: (CATEGORY_PL, "판관비", "급여"),
    18: (CATEGORY_PL, "판관비", "임대용역리스"),
    19: (CATEGORY_PL, "판관비", "세금공과"),
    20: (CATEGORY_PL, "판관비", "기타판관비"),
    21: (CATEGORY_PL, "이익", "당기순이익"),
    22: (CATEGORY_PL, "영업외", "이자비용"),
}

PLAN_START_COL = 3   # C
ACTUAL_START_COL = 19 # S

# 사업계획 파일: 사업계획 시트 KRW 손익 요약 (row 211~235, col 6=1월 ~ col 17=12월)
KSCP_PLAN_MAP = {
    211: (CATEGORY_PL, "매출", "매출액"),
    214: (CATEGORY_PL, "매출원가", "매출원가"),
    215: (CATEGORY_PL, "이익", "매출총이익"),
    217: (CATEGORY_PL, "이익", "영업이익"),
    219: (CATEGORY_MC, "재료비", "재료비계"),
    225: (CATEGORY_MC, "노무비", "노무비계"),
    229: (CATEGORY_PL, "판관비", "판관비계"),
    234: (CATEGORY_PL, "이익", "당기순이익"),
    235: (CATEGORY_PL, "영업외", "이자비용"),
}
KSCP_PLAN_MONTH_COL_START = 6  # F = 1월

class KscpParser(BaseParser):
    def extract(self) -> List[AccountRow]:
        wb = openpyxl.load_workbook(self.filepath, read_only=True, data_only=True)

        if "월종합" not in wb.sheetnames:
            if "사업계획" in wb.sheetnames:
                rows = self._extract_plan(wb)
            else:
                print(f"  [KSCP] 지원 시트 없음 (월종합, 사업계획 모두 없음) — 건너뜀")
                rows = []
            wb.close()
            return rows

        ws = wb["월종합"]
        rows = []

        year = self._detect_year(wb)

        for month_offset in range(12):
            actual_col = ACTUAL_START_COL + month_offset
            month_num = month_offset + 1

            test_val = ws.cell(row=2, column=actual_col).value
            if test_val is None or safe_float(test_val) == 0:
                continue

            ym = f"{year}-{month_num:02d}"

            for row_num, (cat, sub, account) in KSCP_ACCOUNT_MAP.items():
                val = safe_float(ws.cell(row=row_num, column=actual_col).value)
                rows.append(self.make_row(ym, "KSCP", cat, sub, account, "KRW", val))

        wb.close()
        return rows

    def _extract_plan(self, wb) -> List[AccountRow]:
        """사업계획 시트 KRW 손익 요약 (rows 211~235) 에서 월별 데이터 추출"""
        ws = wb["사업계획"]
        year = self._detect_year_plan(ws)
        rows = []
        for month_offset in range(12):
            col = KSCP_PLAN_MONTH_COL_START + month_offset
            month_num = month_offset + 1
            test_val = ws.cell(row=211, column=col).value
            if test_val is None or safe_float(test_val) == 0:
                continue
            ym = f"{year}-{month_num:02d}"
            for row_num, (cat, sub, account) in KSCP_PLAN_MAP.items():
                val = safe_float(ws.cell(row=row_num, column=col).value)
                rows.append(self.make_row(ym, "KSCP", cat, sub, account, "KRW", val, 구분="계획"))
        return rows

    def _detect_year_plan(self, ws) -> str:
        import re
        for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
            for cell in row:
                if cell:
                    m = re.search(r'(\d{4})', str(cell))
                    if m and m.group(1) in ("2024", "2025", "2026", "2027"):
                        return m.group(1)
        return "2026"

    def _detect_year(self, wb) -> str:
        for sheet_name in ["보고용자료", "재무VS관리", "월종합"]:
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
                for cell in row:
                    if cell and "2026" in str(cell):
                        return "2026"
                    if cell and "2025" in str(cell):
                        return "2025"
        return "2026"
