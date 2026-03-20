import openpyxl
import re
from typing import List
from parsers import BaseParser
from schema import AccountRow, safe_float, CATEGORY_PL, CATEGORY_MC

# IS * 시트 (KSCI/KSCE 공용 템플릿): col G = 1월
IS_MONTH_COL_START = 7
KSC_IS_MAP = {
    5:  (CATEGORY_PL, "매출",   "매출액"),
    9:  (CATEGORY_MC, "재료비", "재료비계"),
    21: (CATEGORY_MC, "노무비", "노무비계"),
    35: (CATEGORY_MC, "경비",   "제조경비계"),
    71: (CATEGORY_PL, "판관비", "판관비계"),
}

# PL 시트 (KCTR/KSCE 공용 레이아웃): col 레이블 → 계정
PL_LABEL_MAP = {
    "Ⅰ. 매출액":     (CATEGORY_PL, "매출",    "매출액"),
    "Ⅱ. 매출원가":   (CATEGORY_PL, "매출원가", "매출원가"),
    "Ⅲ. 매출총이익": (CATEGORY_PL, "이익",    "매출총이익"),
    "Ⅳ.판매관리비":  (CATEGORY_PL, "판관비",  "판관비계"),
    "Ⅴ.영업이익":    (CATEGORY_PL, "이익",    "영업이익"),
}
PL_MONTH_COL_START = 6   # F열 = 1월

# 월별실적 시트 (KSCI 스타일): col E = 1월
MONTHLY_MAP = {
    6:  (CATEGORY_PL, "매출",    "매출액"),
    11: (CATEGORY_PL, "매출원가", "매출원가"),
    16: (CATEGORY_PL, "이익",    "매출총이익"),
    18: (CATEGORY_PL, "판관비",  "판관비계"),
}
MONTHLY_COL_START = 5   # E열 = 1월


class KscParser(BaseParser):
    def extract(self) -> List[AccountRow]:
        wb = openpyxl.load_workbook(self.filepath, read_only=True, data_only=True)
        rows = []
        year = self._detect_year(wb)

        # 우선순위 1: 월별실적 시트
        monthly_sheet = next((s for s in wb.sheetnames if "월별실적" in s), None)
        if monthly_sheet:
            rows.extend(self._extract_monthly(wb[monthly_sheet], year))

        # 우선순위 2: PL 시트
        elif "PL" in wb.sheetnames:
            rows.extend(self._extract_pl(wb["PL"], year))

        # 우선순위 3: IS * 시트 (공용 템플릿)
        else:
            is_sheets = [s for s in wb.sheetnames if s.startswith("IS ")]
            if is_sheets:
                rows.extend(self._extract_is_sheets(wb, is_sheets, year))
            else:
                print(f"  [KSC] 지원 시트 없음 (월별실적, PL, IS* 모두 없음) — 건너뜀")

        wb.close()
        return rows

    def _extract_monthly(self, ws, year: str) -> List[AccountRow]:
        rows = []
        for month_offset in range(12):
            col = MONTHLY_COL_START + month_offset
            month_num = month_offset + 1
            test_val = ws.cell(row=6, column=col).value
            if test_val is None or safe_float(test_val) == 0:
                continue
            ym = f"{year}-{month_num:02d}"
            for row_num, (cat, sub, account) in MONTHLY_MAP.items():
                val = safe_float(ws.cell(row=row_num, column=col).value)
                rows.append(self.make_row(ym, "KSC", cat, sub, account, "KRW", val))
            # 영업이익 = 매출총이익 - 판관비
            gross = safe_float(ws.cell(row=16, column=col).value)
            sga = safe_float(ws.cell(row=18, column=col).value)
            rows.append(self.make_row(ym, "KSC", CATEGORY_PL, "이익", "영업이익", "KRW", gross - sga))
        return rows

    def _extract_pl(self, ws, year: str) -> List[AccountRow]:
        label_rows = {}
        for r in range(1, ws.max_row + 1):
            raw = ws.cell(row=r, column=2).value
            if raw is None:
                continue
            key = str(raw).strip()
            if key in PL_LABEL_MAP:
                label_rows[key] = r

        rows = []
        for month_offset in range(12):
            col = PL_MONTH_COL_START + month_offset
            month_num = month_offset + 1
            has_data = any(
                safe_float(ws.cell(row=rn, column=col).value) != 0
                for rn in label_rows.values()
            )
            if not has_data:
                continue
            ym = f"{year}-{month_num:02d}"
            for label, (cat, sub, account) in PL_LABEL_MAP.items():
                if label not in label_rows:
                    continue
                val = safe_float(ws.cell(row=label_rows[label], column=col).value)
                rows.append(self.make_row(ym, "KSC", cat, sub, account, "KRW", val))
        return rows

    def _extract_is_sheets(self, wb, sheet_names: list, year: str) -> List[AccountRow]:
        monthly = {}
        for sname in sheet_names:
            ws = wb[sname]
            for month_offset in range(12):
                col = IS_MONTH_COL_START + month_offset
                if ws.cell(row=5, column=col).value is None:
                    continue
                for row_num in KSC_IS_MAP:
                    key = (month_offset, row_num)
                    monthly[key] = monthly.get(key, 0.0) + safe_float(ws.cell(row=row_num, column=col).value)

        rows = []
        for month_offset in range(12):
            month_num = month_offset + 1
            ym = f"{year}-{month_num:02d}"
            sales = monthly.get((month_offset, 5), 0.0)
            if sales == 0:
                continue
            for row_num, (cat, sub, account) in KSC_IS_MAP.items():
                val = monthly.get((month_offset, row_num), 0.0)
                rows.append(self.make_row(ym, "KSC", cat, sub, account, "KRW", val))
            gross = (monthly.get((month_offset, 5), 0.0)
                     - monthly.get((month_offset, 9), 0.0)
                     - monthly.get((month_offset, 21), 0.0)
                     - monthly.get((month_offset, 35), 0.0))
            rows.append(self.make_row(ym, "KSC", CATEGORY_PL, "이익", "매출총이익", "KRW", gross))
            op = gross - monthly.get((month_offset, 71), 0.0)
            rows.append(self.make_row(ym, "KSC", CATEGORY_PL, "이익", "영업이익", "KRW", op))
        return rows

    def _detect_year(self, wb) -> str:
        for sname in wb.sheetnames:
            ws = wb[sname]
            for r in range(1, 6):
                for c in range(1, 10):
                    val = ws.cell(row=r, column=c).value
                    if not val or not isinstance(val, str):
                        continue
                    m = re.search(r"FY\s*(\d{4})", val)
                    if m:
                        return m.group(1)
                    m = re.search(r"(\d{4})", val)
                    if m and m.group(1) in ("2024", "2025", "2026", "2027"):
                        return m.group(1)
            break  # 첫 번째 시트만 확인
        m = re.search(r"(\d{4})", self.filepath)
        if m:
            return m.group(1)
        return "2026"
