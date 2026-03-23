"""
KSC Refiner v1.3 - Configuration-Driven Parser
설정 기반 범용 파서
"""
import openpyxl
from typing import List, Dict, Optional
import logging

from core.schema import AccountRow, safe_float, detect_year_from_sheet
from parsers.base_parser import BaseParser
from utils.excel_helper import find_sheet, is_valid_data_cell


class ConfigDrivenParser(BaseParser):
    """설정 파일 기반 범용 파서"""

    def extract(self) -> List[AccountRow]:
        """엑셀 파일에서 데이터 추출"""
        try:
            wb = openpyxl.load_workbook(self.filepath, read_only=True, data_only=True)
            rows = []

            # 실적 시트 처리
            actual_sheet = find_sheet(wb, self.config.get("actual_sheets", []))
            if actual_sheet:
                self.log_info(f"실적 시트 발견: {actual_sheet}")
                rows.extend(self._extract_actual(wb, actual_sheet))
            else:
                self.log_warning("실적 시트를 찾을 수 없음")

            # 계획 시트 처리
            plan_sheet = find_sheet(wb, self.config.get("plan_sheets", []))
            if plan_sheet:
                self.log_info(f"계획 시트 발견: {plan_sheet}")
                rows.extend(self._extract_plan(wb, plan_sheet))
            else:
                self.log_warning("계획 시트를 찾을 수 없음")

            wb.close()
            return rows

        except Exception as e:
            self.log_error(f"파일 처리 실패: {e}")
            return []

    def _extract_actual(self, wb: openpyxl.Workbook, sheet_name: str) -> List[AccountRow]:
        """실적 데이터 추출"""
        ws = wb[sheet_name]
        year = detect_year_from_sheet(ws)
        rows = []

        actual_mapping = self.config.get("actual_mapping", {})
        month_start_col = self.config.get("actual_month_start_col", 5)

        # 12개월 순회
        for month_offset in range(12):
            col = month_start_col + month_offset
            month_num = month_offset + 1

            # 해당 월에 데이터가 있는지 확인 (첫 번째 행 체크)
            first_row = min(actual_mapping.keys()) if actual_mapping else 4
            test_val = ws.cell(row=first_row, column=col).value
            if not is_valid_data_cell(test_val):
                continue

            year_month = f"{year}-{month_num:02d}"

            # 매핑된 행들에서 데이터 추출
            for row_num, account_info in actual_mapping.items():
                category, subcategory, account = account_info
                cell_value = ws.cell(row=row_num, column=col).value
                amount = safe_float(cell_value)

                # 0이 아닌 값만 추가 (선택적)
                # if amount != 0:
                rows.append(self.make_row(
                    year_month=year_month,
                    category=category,
                    subcategory=subcategory,
                    account=account,
                    local_amount=amount,
                    data_type="실적"
                ))

        return rows

    def _extract_plan(self, wb: openpyxl.Workbook, sheet_name: str) -> List[AccountRow]:
        """계획 데이터 추출"""
        ws = wb[sheet_name]
        year = detect_year_from_sheet(ws)
        rows = []

        plan_mapping = self.config.get("plan_mapping", {})
        month_start_col = self.config.get("plan_month_start_col", 5)

        # 12개월 순회
        for month_offset in range(12):
            col = month_start_col + month_offset
            month_num = month_offset + 1

            # 해당 월에 데이터가 있는지 확인
            first_row = min(plan_mapping.keys()) if plan_mapping else 4
            test_val = ws.cell(row=first_row, column=col).value
            if not is_valid_data_cell(test_val):
                continue

            year_month = f"{year}-{month_num:02d}"

            # 매핑된 행들에서 데이터 추출
            for row_num, account_info in plan_mapping.items():
                category, subcategory, account = account_info
                cell_value = ws.cell(row=row_num, column=col).value
                amount = safe_float(cell_value)

                rows.append(self.make_row(
                    year_month=year_month,
                    category=category,
                    subcategory=subcategory,
                    account=account,
                    local_amount=amount,
                    data_type="계획"
                ))

        return rows
