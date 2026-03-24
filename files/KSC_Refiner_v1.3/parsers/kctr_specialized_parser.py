"""
KSC Refiner v1.3 - KCTR Specialized Parser
KCTR 법인 전용 파서 (★PL(Report) 시트 처리)
"""
import openpyxl
from typing import List

from core.schema import AccountRow, safe_float, detect_year_from_sheet
from parsers.config_driven_parser import ConfigDrivenParser
from utils.excel_helper import find_sheet


class KctrSpecializedParser(ConfigDrivenParser):
    """KCTR 특수 케이스 처리 파서"""

    def extract(self) -> List[AccountRow]:
        """KCTR 파일에서 데이터 추출"""
        try:
            wb = openpyxl.load_workbook(self.filepath, read_only=True, data_only=True)
            rows = []

            # ★PL(Report) 시트 처리 (특수 케이스)
            if "★PL(Report)" in wb.sheetnames:
                self.log_info("★PL(Report) 시트 감지 - 레이블 매칭 모드")
                rows.extend(self._extract_pl_report(wb))
            else:
                # 일반 시트 처리 (설정 기반)
                actual_sheet = find_sheet(wb, self.config.get("actual_sheets", []))
                if actual_sheet and actual_sheet != "★PL(Report)":
                    self.log_info(f"실적 시트 발견: {actual_sheet}")
                    rows.extend(self._extract_actual(wb, actual_sheet))

            # 계획 시트 처리
            plan_sheet = find_sheet(wb, self.config.get("plan_sheets", []))
            if plan_sheet:
                self.log_info(f"계획 시트 발견: {plan_sheet}")
                rows.extend(self._extract_plan(wb, plan_sheet))

            wb.close()
            return rows

        except Exception as e:
            self.log_error(f"파일 처리 실패: {e}")
            return []

    def _extract_pl_report(self, wb: openpyxl.Workbook) -> List[AccountRow]:
        """
        ★PL(Report) 시트에서 누적 실적 추출
        레이블 기반 매칭 사용
        """
        ws = wb["★PL(Report)"]
        year = detect_year_from_sheet(ws)
        rows = []

        # Row 4, Col 3에서 월 정보 읽기
        month_val = ws.cell(row=4, column=3).value
        if month_val is None:
            self.log_warning("★PL(Report): 월 정보 없음")
            return rows

        try:
            month_num = int(safe_float(month_val))
            if month_num <= 0 or month_num > 12:
                self.log_warning(f"★PL(Report): 유효하지 않은 월 정보: {month_num}")
                return rows
        except ValueError:
            self.log_warning(f"★PL(Report): 월 정보 파싱 실패: {month_val}")
            return rows

        year_month = f"{year}-{month_num:02d}"

        # 레이블 패턴 매칭
        label_patterns = self.config.get("label_patterns", {})

        # Row 6~70 범위에서 레이블 검색
        for row_num in range(6, min(ws.max_row + 1, 71)):
            label_cell = ws.cell(row=row_num, column=2).value
            if not label_cell:
                continue

            label_str = str(label_cell).strip().upper()
            actual_val = safe_float(ws.cell(row=row_num, column=5).value)  # Col 5 = Accumulated Actual

            # 각 계정과목 패턴과 매칭
            matched = False
            for account, patterns in label_patterns.items():
                for pattern in patterns:
                    if pattern.upper() in label_str:
                        # 계정과목에 따라 카테고리/서브카테고리 결정
                        category, subcategory = self._get_category_for_account(account)
                        rows.append(self.make_row(
                            year_month=year_month,
                            category=category,
                            subcategory=subcategory,
                            account=account,
                            local_amount=actual_val,
                            data_type="실적"
                        ))
                        matched = True
                        break
                if matched:
                    break

        self.log_info(f"★PL(Report): {len(rows)}건 추출 ({year_month})")
        return rows

    def _get_category_for_account(self, account: str) -> tuple:
        """
        계정과목에 따라 대분류/중분류 반환

        Args:
            account: 계정과목

        Returns:
            (대분류, 중분류)
        """
        # PL 계정
        pl_accounts = {
            "매출액": ("PL", "매출"),
            "상품매출": ("PL", "매출"),
            "매출원가": ("PL", "매출원가"),
            "매출총이익": ("PL", "이익"),
            "판관비계": ("PL", "판관비"),
            "영업이익": ("PL", "이익"),
        }
        # MC 계정
        mc_accounts = {
            "재료비계": ("MC", "재료비"),
            "노무비계": ("MC", "노무비"),
            "제조경비계": ("MC", "경비"),
        }

        if account in pl_accounts:
            return pl_accounts[account]
        if account in mc_accounts:
            return mc_accounts[account]

        # 기본값
        return ("PL", "기타")
