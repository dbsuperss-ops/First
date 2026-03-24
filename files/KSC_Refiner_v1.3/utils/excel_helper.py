"""
KSC Refiner v1.3 - Excel Helper Utilities
엑셀 파일 처리 유틸리티
"""
import openpyxl
from typing import Optional, List


def find_sheet(workbook: openpyxl.Workbook, sheet_names: List[str]) -> Optional[str]:
    """
    워크북에서 주어진 시트명 목록 중 하나를 찾음

    Args:
        workbook: openpyxl Workbook 객체
        sheet_names: 찾을 시트명 리스트

    Returns:
        찾은 시트명 또는 None
    """
    for name in sheet_names:
        if name in workbook.sheetnames:
            return name
    return None


def find_row_by_label(worksheet, label_keywords: List[str],
                      search_col: int = 2, max_row: int = 100) -> Optional[int]:
    """
    시트에서 특정 레이블을 포함하는 행 번호 찾기

    Args:
        worksheet: openpyxl Worksheet 객체
        label_keywords: 찾을 키워드 리스트 (OR 조건)
        search_col: 검색할 열 번호
        max_row: 최대 검색 행

    Returns:
        행 번호 또는 None
    """
    for row_num in range(1, min(worksheet.max_row + 1, max_row + 1)):
        cell_value = worksheet.cell(row=row_num, column=search_col).value
        if cell_value:
            cell_str = str(cell_value).strip()
            for keyword in label_keywords:
                if keyword in cell_str:
                    return row_num
    return None


def get_column_letter_from_index(col_idx: int) -> str:
    """
    열 인덱스를 엑셀 열 문자로 변환

    Args:
        col_idx: 1-based 열 인덱스

    Returns:
        엑셀 열 문자 (A, B, ..., AA, AB, ...)
    """
    return openpyxl.utils.get_column_letter(col_idx)


def is_valid_data_cell(value) -> bool:
    """
    셀 값이 유효한 데이터인지 확인

    Args:
        value: 셀 값

    Returns:
        True if valid data
    """
    if value is None:
        return False
    if isinstance(value, str) and value.strip() in ("", "-", "N/A"):
        return False
    try:
        float_val = float(value)
        return float_val != 0
    except (ValueError, TypeError):
        return False
