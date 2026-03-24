"""
KSC Refiner v1.3 - Output Writer
통합 DB 엑셀 파일 생성
"""
from typing import List
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import logging

from core.schema import AccountRow, MASTER_COLUMNS


class OutputWriter:
    """통합 DB 엑셀 파일 작성기"""

    def __init__(self):
        self.logger = logging.getLogger("OutputWriter")

    def write_consolidated(self, all_rows: List[AccountRow], output_path: str):
        """
        통합 DB 엑셀 파일 생성

        Args:
            all_rows: 전체 데이터 행 리스트
            output_path: 출력 파일 경로
        """
        self.logger.info(f"통합 DB 작성 시작: {len(all_rows)}건")

        wb = Workbook()
        ws = wb.active
        ws.title = "통합_DB"

        # 헤더 스타일
        header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill("solid", fgColor="2F5496")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9"),
        )

        # 헤더 작성
        for col_idx, col_name in enumerate(MASTER_COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        # 데이터 작성
        num_fmt_local = '#,##0.00'
        num_fmt_krw = '#,##0'
        num_fmt_rate = '#,##0.00'

        for row_idx, account_row in enumerate(all_rows, 2):
            data = account_row.to_list()
            for col_idx, val in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = thin_border
                cell.font = Font(name="Arial", size=10)

                # 숫자 포맷 적용
                if col_idx == 8:  # 현지금액
                    cell.number_format = num_fmt_local
                elif col_idx == 9:  # 적용환율
                    cell.number_format = num_fmt_rate
                elif col_idx == 10:  # KRW금액
                    cell.number_format = num_fmt_krw

        # 열 너비 조정
        col_widths = [12, 10, 8, 6, 10, 16, 8, 18, 10, 18]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

        # 필터 및 틀 고정
        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = "A2"

        # 법인별 요약 시트 추가
        self._create_summary_sheet(wb, all_rows)

        # 경영보고서 시트 추가
        self._create_executive_report(wb, all_rows)

        # 파일 저장
        wb.save(output_path)
        self.logger.info(f"통합 DB 저장 완료: {output_path}")

        # 요약 출력
        companies = sorted(set(r.법인코드 for r in all_rows))
        self.logger.info(f"총 {len(all_rows)}건, {len(companies)}개 법인")
        self.logger.info(f"포함 시트: 통합_DB, 법인별_요약, 경영보고서")

    def _create_summary_sheet(self, wb: Workbook, all_rows: List[AccountRow]):
        """법인별 요약 시트 생성"""
        ws = wb.create_sheet("법인별_요약")

        companies = sorted(set(r.법인코드 for r in all_rows))
        actual_rows = [r for r in all_rows if r.데이터타입 == "실적"]
        periods = sorted(set(r.귀속연월 for r in actual_rows)) or sorted(set(r.귀속연월 for r in all_rows))

        key_accounts = ["매출액", "매출원가", "매출총이익", "판관비계", "영업이익"]

        # 헤더
        header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        header_fill = PatternFill("solid", fgColor="2F5496")
        thin_border = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9"),
        )

        headers = ["귀속연월", "계정과목"] + companies + ["합계"]
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        # 데이터
        row_idx = 2
        for period in periods:
            for account in key_accounts:
                ws.cell(row=row_idx, column=1, value=period).border = thin_border
                ws.cell(row=row_idx, column=2, value=account).border = thin_border

                total = 0
                for comp_idx, comp in enumerate(companies, 3):
                    val = sum(
                        r.KRW금액 for r in actual_rows
                        if r.귀속연월 == period and r.법인코드 == comp and r.계정과목 == account
                    )
                    cell = ws.cell(row=row_idx, column=comp_idx, value=val)
                    cell.number_format = '#,##0'
                    cell.border = thin_border
                    total += val

                cell = ws.cell(row=row_idx, column=len(companies) + 3, value=total)
                cell.number_format = '#,##0'
                cell.border = thin_border
                row_idx += 1

        # 열 너비
        col_widths = [12, 14] + [16] * (len(companies) + 1)
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

        ws.freeze_panes = "C2"

    def _create_executive_report(self, wb: Workbook, all_rows: List[AccountRow]):
        """경영보고서 시트 생성"""
        ws = wb.create_sheet("경영보고서")

        companies = sorted(set(r.법인코드 for r in all_rows))
        actual_rows = [r for r in all_rows if r.데이터타입 == "실적"]
        plan_rows = [r for r in all_rows if r.데이터타입 == "계획"]

        actual_periods = sorted(set(r.귀속연월 for r in actual_rows))
        latest_period = actual_periods[-1] if actual_periods else (
            sorted(set(r.귀속연월 for r in all_rows))[-1] if all_rows else "N/A"
        )

        # 스타일
        hdr_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        hdr_fill = PatternFill("solid", fgColor="2F5496")
        sub_hdr_fill = PatternFill("solid", fgColor="D6E4F0")
        sub_hdr_font = Font(name="Arial", bold=True, size=10)
        profit_fill = PatternFill("solid", fgColor="E2EFDA")
        loss_font = Font(name="Arial", size=10, color="FF0000")
        normal_font = Font(name="Arial", size=10)
        thin_b = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9"),
        )

        # 제목
        period_range_start = actual_periods[0] if actual_periods else latest_period
        ws.cell(row=1, column=1, value="경신그룹 해외법인 경영실적 종합").font = Font(name="Arial", bold=True, size=14)
        ws.cell(row=2, column=1, value=f"당월(실적): {latest_period} | 누적: {period_range_start} ~ {latest_period} | 단위: 백만원(KRW)").font = Font(name="Arial", size=10, color="808080")

        # 헤더 (4행)
        row = 4
        n = len(companies)
        headers = (
            ["구분", "계정과목"]
            + [f"{c}\n실적당월" for c in companies]
            + [f"{c}\n계획당월" for c in companies]
            + [f"{c}\n실적누계" for c in companies]
            + [f"{c}\n계획누계" for c in companies]
            + ["합계\n(실적누계)"]
        )
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=ci, value=h)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_b

        row += 1

        # 데이터 작성 함수
        def get_val(corp, acct, period=None, dtype=None):
            filtered = all_rows
            if dtype:
                filtered = [r for r in filtered if r.데이터타입 == dtype]
            if period:
                filtered = [r for r in filtered if r.귀속연월 == period]
            return sum(r.KRW금액 for r in filtered if r.법인코드 == corp and r.계정과목 == acct)

        def get_val_ytd(corp, acct, up_to_period, dtype=None):
            filtered = all_rows
            if dtype:
                filtered = [r for r in filtered if r.데이터타입 == dtype]
            return sum(r.KRW금액 for r in filtered
                       if r.법인코드 == corp and r.계정과목 == acct and r.귀속연월 <= up_to_period)

        def write_account_row(ws, row, section, acct, companies, latest, is_key=False):
            ws.cell(row=row, column=1, value=section).border = thin_b
            cell_acct = ws.cell(row=row, column=2, value=acct)
            cell_acct.border = thin_b
            if is_key:
                cell_acct.font = sub_hdr_font

            total_actual_cum = 0
            for ci, comp in enumerate(companies):
                act_m = get_val(corp=comp, acct=acct, period=latest, dtype="실적") / 1_000_000
                plan_m = get_val(corp=comp, acct=acct, period=latest, dtype="계획") / 1_000_000
                act_c = get_val_ytd(comp, acct, latest, dtype="실적") / 1_000_000
                plan_c = get_val_ytd(comp, acct, latest, dtype="계획") / 1_000_000

                def write_cell(col, val):
                    c = ws.cell(row=row, column=col, value=round(val, 0))
                    c.number_format = '#,##0'
                    c.border = thin_b
                    c.font = loss_font if val < 0 else normal_font

                write_cell(3 + ci, act_m)
                write_cell(3 + n + ci, plan_m)
                write_cell(3 + 2 * n + ci, act_c)
                write_cell(3 + 3 * n + ci, plan_c)

                total_actual_cum += act_c

            t_cell = ws.cell(row=row, column=3 + 4 * n, value=round(total_actual_cum, 0))
            t_cell.number_format = '#,##0'
            t_cell.border = thin_b
            t_cell.font = Font(name="Arial", bold=True, size=10, color="FF0000" if total_actual_cum < 0 else "000000")

            if acct in ("매출총이익", "영업이익"):
                for c in range(1, 3 + 4 * n + 1):
                    ws.cell(row=row, column=c).fill = profit_fill

        # PL 계정
        total_cols = 3 + 4 * n
        pl_accounts = [
            ("매출", ["매출액", "상품매출", "제품매출", "전선매출", "전장매출", "기타매출"]),
            ("매출원가", ["매출원가", "제품매출원가", "상품매출원가"]),
            ("이익", ["매출총이익"]),
            ("판관비", ["판관비계"]),
            ("이익", ["영업이익"]),
        ]

        ws.cell(row=row, column=1, value="【 PL 】").font = sub_hdr_font
        ws.cell(row=row, column=1).fill = sub_hdr_fill
        for ci in range(2, total_cols + 1):
            ws.cell(row=row, column=ci).fill = sub_hdr_fill
            ws.cell(row=row, column=ci).border = thin_b
        row += 1

        for section, accts in pl_accounts:
            for acct in accts:
                has_data = any(
                    get_val(corp=c, acct=acct, dtype="실적") != 0 or get_val(corp=c, acct=acct, dtype="계획") != 0
                    for c in companies
                )
                if has_data:
                    is_key = acct in ("매출액", "매출원가", "매출총이익", "판관비계", "영업이익")
                    write_account_row(ws, row, section, acct, companies, latest_period, is_key)
                    row += 1

        # 제조원가
        row += 1
        mc_accounts = [
            ("재료비", ["재료비계"]),
            ("노무비", ["노무비계"]),
            ("경비", ["제조경비계", "기계경비"]),
        ]

        ws.cell(row=row, column=1, value="【 제조원가 】").font = sub_hdr_font
        ws.cell(row=row, column=1).fill = sub_hdr_fill
        for ci in range(2, total_cols + 1):
            ws.cell(row=row, column=ci).fill = sub_hdr_fill
            ws.cell(row=row, column=ci).border = thin_b
        row += 1

        for section, accts in mc_accounts:
            for acct in accts:
                has_data = any(
                    get_val(corp=c, acct=acct, dtype="실적") != 0 or get_val(corp=c, acct=acct, dtype="계획") != 0
                    for c in companies
                )
                if has_data:
                    write_account_row(ws, row, section, acct, companies, latest_period)
                    row += 1

        # 이익률
        row += 1
        ws.cell(row=row, column=1, value="【 이익률 (%) 】").font = sub_hdr_font
        ws.cell(row=row, column=1).fill = sub_hdr_fill
        for ci in range(2, total_cols + 1):
            ws.cell(row=row, column=ci).fill = sub_hdr_fill
        row += 1

        for label in ["매출총이익률", "영업이익률"]:
            ws.cell(row=row, column=1, value="이익률").border = thin_b
            ws.cell(row=row, column=2, value=label).border = thin_b
            ws.cell(row=row, column=2).font = sub_hdr_font

            acct = "매출총이익" if "총" in label else "영업이익"
            for ci, comp in enumerate(companies):
                rev_m_a = get_val(corp=comp, acct="매출액", period=latest_period, dtype="실적")
                prof_m_a = get_val(corp=comp, acct=acct, period=latest_period, dtype="실적")
                rev_m_p = get_val(corp=comp, acct="매출액", period=latest_period, dtype="계획")
                prof_m_p = get_val(corp=comp, acct=acct, period=latest_period, dtype="계획")
                rev_c_a = get_val_ytd(comp, "매출액", latest_period, dtype="실적")
                prof_c_a = get_val_ytd(comp, acct, latest_period, dtype="실적")
                rev_c_p = get_val_ytd(comp, "매출액", latest_period, dtype="계획")
                prof_c_p = get_val_ytd(comp, acct, latest_period, dtype="계획")

                def pct_cell(col, num, den):
                    c = ws.cell(row=row, column=col, value=num / den if den else 0)
                    c.number_format = '0.0%'
                    c.border = thin_b

                pct_cell(3 + ci, prof_m_a, rev_m_a)
                pct_cell(3 + n + ci, prof_m_p, rev_m_p)
                pct_cell(3 + 2 * n + ci, prof_c_a, rev_c_a)
                pct_cell(3 + 3 * n + ci, prof_c_p, rev_c_p)

            row += 1

        # 열 너비
        col_widths = [10, 16] + [13] * (4 * n + 1)
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[ws.cell(row=4, column=i).column_letter].width = w

        ws.freeze_panes = "C5"
        ws.sheet_properties.tabColor = "2F5496"
