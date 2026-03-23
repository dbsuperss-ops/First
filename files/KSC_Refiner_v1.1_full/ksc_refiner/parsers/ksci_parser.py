import openpyxl
import re
from typing import List
from parsers import BaseParser
from schema import AccountRow, safe_float, CATEGORY_PL, CATEGORY_MC

KSCI_PL_MAP = {
    6:  (CATEGORY_PL, "매출", "매출액"),
    7:  (CATEGORY_PL, "매출", "전선매출"),
    9:  (CATEGORY_PL, "매출", "전장매출"),
    11: (CATEGORY_PL, "매출원가", "매출원가"),
    12: (CATEGORY_PL, "매출원가", "제품매출원가"),
    13: (CATEGORY_PL, "매출원가", "상품매출원가"),
    14: (CATEGORY_PL, "매출원가", "기타매출원가"),
    16: (CATEGORY_PL, "이익", "매출총이익"),
    18: (CATEGORY_PL, "판관비", "판관비계"),
}

MONTH_COL_START = 5  # E = 1월

# 사업계획 파일: IS DGO + IS 전장 시트 기반 (col 7~18 = Jan~Dec)
KSCI_PLAN_IS_MAP = {
    5:  (CATEGORY_PL, "매출", "매출액"),
    9:  (CATEGORY_MC, "재료비", "재료비계"),
    21: (CATEGORY_MC, "노무비", "노무비계"),
    35: (CATEGORY_MC, "경비", "제조경비계"),
    71: (CATEGORY_PL, "판관비", "판관비계"),
}

IS_PLAN_MONTH_COL_START = 7  # col 7 = Jan(1월)
IS_PLAN_SHEETS = ["IS DGO", "IS 전장"]

class KsciParser(BaseParser):
    def extract(self) -> List[AccountRow]:
        wb = openpyxl.load_workbook(self.filepath, read_only=True, data_only=True)
        rows = []

        target_sheet = None
        for name in wb.sheetnames:
            if "월별실적" in name:
                target_sheet = name
                break

        if not target_sheet:
            plan_sheets = [s for s in IS_PLAN_SHEETS if s in wb.sheetnames]
            if plan_sheets:
                rows.extend(self._extract_plan(wb, plan_sheets))
            else:
                print(f"  [KSCI] 지원 시트 없음 (월별실적, IS DGO/IS 전장 모두 없음) — 건너뜀")
            wb.close()
            return rows

        ws = wb[target_sheet]
        year = self._detect_year(ws)

        for month_offset in range(12):
            col = MONTH_COL_START + month_offset
            month_num = month_offset + 1

            test_val = ws.cell(row=6, column=col).value
            if test_val is None or safe_float(test_val) == 0:
                continue

            ym = f"{year}-{month_num:02d}"
            for row_num, (cat, sub, account) in KSCI_PL_MAP.items():
                val = safe_float(ws.cell(row=row_num, column=col).value)
                rows.append(self.make_row(ym, "KSCI", cat, sub, account, "USD", val))

            op_profit = safe_float(ws.cell(row=16, column=col).value) - safe_float(ws.cell(row=18, column=col).value)
            rows.append(self.make_row(ym, "KSCI", CATEGORY_PL, "이익", "영업이익", "USD", op_profit))

        rows.extend(self._extract_mc_summary(wb, year))

        wb.close()
        return rows

    def _extract_plan(self, wb, plan_sheets: list) -> List[AccountRow]:
        """IS DGO + IS 전장 시트를 합산하여 사업계획 월별 데이터 추출"""
        year = "2026"
        # 시트에서 연도 감지
        for shname in plan_sheets:
            ws = wb[shname]
            for row in ws.iter_rows(min_row=1, max_row=3, values_only=True):
                for cell in row:
                    if cell:
                        import re as _re
                        m = _re.search(r'(\d{4})', str(cell))
                        if m and m.group(1) in ("2024", "2025", "2026", "2027"):
                            year = m.group(1)
                            break

        # (month_offset, row_num) → 합산값
        monthly = {}
        for shname in plan_sheets:
            ws = wb[shname]
            for month_offset in range(12):
                col = IS_PLAN_MONTH_COL_START + month_offset
                test_val = ws.cell(row=5, column=col).value
                if test_val is None:
                    continue
                for row_num in KSCI_PLAN_IS_MAP:
                    key = (month_offset, row_num)
                    monthly[key] = monthly.get(key, 0.0) + safe_float(ws.cell(row=row_num, column=col).value)

        rows = []
        for month_offset in range(12):
            month_num = month_offset + 1
            ym = f"{year}-{month_num:02d}"
            sales = monthly.get((month_offset, 5), 0.0)
            if sales == 0:
                continue
            for row_num, (cat, sub, account) in KSCI_PLAN_IS_MAP.items():
                val = monthly.get((month_offset, row_num), 0.0)
                rows.append(self.make_row(ym, "KSCI", cat, sub, account, "USD", val, data_type="계획"))
            # 영업이익 = 매출 - 재료비 - 노무비 - 경비 - 판관비
            op = (monthly.get((month_offset, 5), 0.0)
                  - monthly.get((month_offset, 9), 0.0)
                  - monthly.get((month_offset, 21), 0.0)
                  - monthly.get((month_offset, 35), 0.0)
                  - monthly.get((month_offset, 71), 0.0))
            rows.append(self.make_row(ym, "KSCI", CATEGORY_PL, "이익", "영업이익", "USD", op, data_type="계획"))
        return rows

    def _extract_mc_summary(self, wb, year) -> List[AccountRow]:
        summary_sheet = None
        for name in wb.sheetnames:
            if "요약" in name or "경신전선" in name:
                summary_sheet = name
                break
        if not summary_sheet:
            return []

        ws = wb[summary_sheet]
        mc_rows = []

        mc_keywords = {
            "재료비": (CATEGORY_MC, "재료비", "재료비계"),
            "노무비": (CATEGORY_MC, "노무비", "노무비계"),
            "경비":   (CATEGORY_MC, "경비", "제조경비계"),
        }

        month_col = None
        month_num = None
        for row in ws.iter_rows(min_row=4, max_row=6, values_only=False):
            for cell in row:
                if cell.value and isinstance(cell.value, str) and "매출액" in str(cell.value):
                    for adj_col in range(cell.column + 1, cell.column + 10):
                        hdr = ws.cell(row=cell.row - 1, column=adj_col).value
                        if hdr and "실적" in str(hdr) and "누" not in str(hdr):
                            month_col = adj_col
                            break

        if month_col is None:
            return []

        for r in range(20, min(ws.max_row + 1, 60)):
            label = ws.cell(row=r, column=3).value
            if not label:
                continue
            label_str = str(label).strip()
            for kw, mapping in mc_keywords.items():
                if kw in label_str and "Ⅰ" not in label_str and "Ⅱ" not in label_str:
                    continue
                if kw in label_str:
                    val = safe_float(ws.cell(row=r, column=month_col).value)
                    if val == 0:
                        val = safe_float(ws.cell(row=r, column=7).value)
                    if val != 0:
                        ym = f"{year}-01"
                        mc_rows.append(self.make_row(ym, "KSCI", mapping[0], mapping[1], mapping[2], "USD", val))
                    break

        return mc_rows

    def _detect_year(self, ws) -> str:
        for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
            for cell in row:
                if cell:
                    m = re.search(r'FY\s*(\d{4})', str(cell))
                    if m:
                        return m.group(1)
                    m = re.search(r'(\d{4})', str(cell))
                    if m and m.group(1) in ("2024", "2025", "2026", "2027"):
                        return m.group(1)
        return "2025"
