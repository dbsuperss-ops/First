import openpyxl
import re
from typing import List
from parsers import BaseParser
from schema import AccountRow, safe_float, CATEGORY_PL, CATEGORY_MC

KSCCZ_PL_MAP = {
    3:  (CATEGORY_PL, "매출", "매출액"),
    4:  (CATEGORY_PL, "매출", "제품매출"),
    5:  (CATEGORY_PL, "매출", "상품매출"),
    6:  (CATEGORY_PL, "매출", "기타매출"),
    10: (CATEGORY_PL, "매출원가", "매출원가"),
    11: (CATEGORY_PL, "매출원가", "제품매출원가"),
    12: (CATEGORY_PL, "매출원가", "상품매출원가"),
    16: (CATEGORY_PL, "이익", "매출총이익"),
    18: (CATEGORY_PL, "판관비", "판관비계"),
    19: (CATEGORY_PL, "판관비", "급여"),
    20: (CATEGORY_PL, "판관비", "복리후생비"),
    21: (CATEGORY_PL, "판관비", "운반비"),
    22: (CATEGORY_PL, "판관비", "여비교통비"),
    23: (CATEGORY_PL, "판관비", "지급수수료"),
    24: (CATEGORY_PL, "판관비", "통신비"),
    25: (CATEGORY_PL, "판관비", "접대비"),
    26: (CATEGORY_PL, "판관비", "감가상각비"),
}

KSCCZ_MC_MAP = {
    3:  (CATEGORY_MC, "재료비", "재료비계"),
    4:  (CATEGORY_MC, "재료비", "기초재고"),
    5:  (CATEGORY_MC, "재료비", "당기매입액"),
    7:  (CATEGORY_MC, "재료비", "기말재고액"),
    8:  (CATEGORY_MC, "노무비", "노무비계"),
    10: (CATEGORY_MC, "경비", "제조경비계"),
    11: (CATEGORY_MC, "경비", "복리후생비"),
    15: (CATEGORY_MC, "경비", "수도광열비"),
    17: (CATEGORY_MC, "경비", "감가상각비"),
    19: (CATEGORY_MC, "경비", "수선비"),
    20: (CATEGORY_MC, "경비", "보험료"),
}

ACTUAL_COL = 7  # G column = 실적

# 사업계획 파일: 손익계획 시트 (col 5=1월 ~ col 16=12월)
KSCCZ_PLAN_PL_MAP = {
    4:  (CATEGORY_PL, "매출",    "매출액"),
    16: (CATEGORY_PL, "매출원가", "매출원가"),
    22: (CATEGORY_PL, "이익",    "매출총이익"),
    24: (CATEGORY_PL, "판관비",  "판관비계"),
}
# 사업계획 파일: 제조원가계획 시트 (col 5=1월 ~ col 16=12월)
KSCCZ_PLAN_MC_MAP = {
    4:  (CATEGORY_MC, "재료비", "재료비계"),
    11: (CATEGORY_MC, "노무비", "노무비계"),
    16: (CATEGORY_MC, "경비",  "제조경비계"),
}
KSCCZ_PLAN_MONTH_COL_START = 5  # E column = 1월

class KscczParser(BaseParser):
    def extract(self) -> List[AccountRow]:
        wb = openpyxl.load_workbook(self.filepath, read_only=True, data_only=True)
        rows = []

        if "★손익계산서" not in wb.sheetnames:
            if "손익계획" in wb.sheetnames:
                rows = self._extract_plan(wb)
            else:
                print(f"  [KSCCZ] 지원 시트 없음 (★손익계산서, 손익계획 모두 없음) — 건너뜀")
            wb.close()
            return rows

        ws_pl = wb["★손익계산서"]
        month_num = self._detect_month(ws_pl)
        year = self._detect_year(ws_pl)
        ym = f"{year}-{month_num:02d}"

        for row_num, (cat, sub, account) in KSCCZ_PL_MAP.items():
            val = safe_float(ws_pl.cell(row=row_num, column=ACTUAL_COL).value)
            rows.append(self.make_row(ym, "KSCCZ", cat, sub, account, "RMB", val))

        op_profit = (
            safe_float(ws_pl.cell(row=16, column=ACTUAL_COL).value) -
            safe_float(ws_pl.cell(row=18, column=ACTUAL_COL).value)
        )
        rows.append(self.make_row(ym, "KSCCZ", CATEGORY_PL, "이익", "영업이익", "RMB", op_profit))

        if "★제조원가명세서" in wb.sheetnames:
            ws_mc = wb["★제조원가명세서"]
            for row_num, (cat, sub, account) in KSCCZ_MC_MAP.items():
                val = safe_float(ws_mc.cell(row=row_num, column=ACTUAL_COL).value)
                rows.append(self.make_row(ym, "KSCCZ", cat, sub, account, "RMB", val))

        wb.close()
        return rows

    def _extract_plan(self, wb) -> List[AccountRow]:
        """사업계획 파일: 손익계획 + 제조원가계획 시트에서 월별 데이터 추출"""
        ws_pl = wb["손익계획"]
        year = self._detect_year(ws_pl)
        rows = []

        for month_offset in range(12):
            col = KSCCZ_PLAN_MONTH_COL_START + month_offset
            month_num = month_offset + 1
            test_val = ws_pl.cell(row=4, column=col).value
            if test_val is None or safe_float(test_val) == 0:
                continue
            ym = f"{year}-{month_num:02d}"

            for row_num, (cat, sub, account) in KSCCZ_PLAN_PL_MAP.items():
                val = safe_float(ws_pl.cell(row=row_num, column=col).value)
                rows.append(self.make_row(ym, "KSCCZ", cat, sub, account, "RMB", val, data_type="계획"))

            # 영업이익: 시트에 직접 값이 있는 경우 우선 사용, 없으면 매출총이익 - 판관비계 계산
            op_profit = self._find_plan_op_profit(ws_pl, col)
            rows.append(self.make_row(ym, "KSCCZ", CATEGORY_PL, "이익", "영업이익", "RMB", op_profit, data_type="계획"))

            if "제조원가계획" in wb.sheetnames:
                ws_mc = wb["제조원가계획"]
                for row_num, (cat, sub, account) in KSCCZ_PLAN_MC_MAP.items():
                    val = safe_float(ws_mc.cell(row=row_num, column=col).value)
                    rows.append(self.make_row(ym, "KSCCZ", cat, sub, account, "RMB", val, data_type="계획"))

        return rows

    def _find_plan_op_profit(self, ws_pl, col) -> float:
        """손익계획 시트에서 영업이익 값을 찾음.
        1) '营业利润' 또는 '영업이익' 레이블 행을 탐색해 직접 읽음.
        2) 못 찾으면 row22(매출총이익) - row24(판관비계) 계산값 반환."""
        search_keywords = ["营业利润", "영업이익", "营业利润合计"]
        for r in range(1, min(ws_pl.max_row + 1, 60)):
            label = ws_pl.cell(row=r, column=1).value or ws_pl.cell(row=r, column=2).value
            if label and any(kw in str(label) for kw in search_keywords):
                val = safe_float(ws_pl.cell(row=r, column=col).value)
                print(f"    📌 영업이익 행 발견 (row {r}): {val}")
                return val
        # 레이블 미발견 → 계산
        gp = safe_float(ws_pl.cell(row=22, column=col).value)
        sga = safe_float(ws_pl.cell(row=24, column=col).value)
        print(f"    ⚠️ 영업이익 행 미발견 → 매출총이익({gp}) - 판관비({sga}) = {gp - sga}")
        return gp - sga

    def _detect_month(self, ws) -> int:
        val = ws.cell(row=1, column=2).value
        if val and isinstance(val, (int, float)):
            return int(val)
        for row in ws.iter_rows(min_row=1, max_row=3, values_only=True):
            for cell in row:
                if cell:
                    m = re.search(r'(\d{1,2})\s*월', str(cell))
                    if m:
                        return int(m.group(1))
        return 1

    def _detect_year(self, ws) -> str:
        for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
            for cell in row:
                if cell:
                    m = re.search(r'(\d{4})', str(cell))
                    if m and m.group(1) in ("2024", "2025", "2026", "2027"):
                        return m.group(1)
        return "2026"
