import openpyxl
from typing import List
from parsers import BaseParser
from schema import AccountRow, safe_float, CATEGORY_PL, CATEGORY_MC

# KSCE 손익계산서 매핑
KSCE_PL_MAP = {
    4:  (CATEGORY_PL, "매출", "매출액"),
    5:  (CATEGORY_PL, "매출", "상품매출"),
    15: (CATEGORY_PL, "매출", "기타매출"),
    23: (CATEGORY_PL, "매출원가", "매출원가"),
    25: (CATEGORY_PL, "매출원가", "상품매출원가"),
    37: (CATEGORY_PL, "이익", "매출총이익"),
    39: (CATEGORY_PL, "판관비", "판관비계"),
    65: (CATEGORY_PL, "이익", "영업이익"),
}

# KSCE 제조원가 매핑
KSCE_MC_MAP = {
    62: (CATEGORY_MC, "재료비", "재료비계"),
    63: (CATEGORY_MC, "노무비", "노무비계"),
    64: (CATEGORY_MC, "경비", "제조경비계"),
}

MONTH_COL_START = 5  # E column = 1월

class KsceParser(BaseParser):
    def extract(self) -> List[AccountRow]:
        wb = openpyxl.load_workbook(self.filepath, read_only=True, data_only=True)
        rows = []

        # 시트 찾기 (KSCE 실적, Summary 등)
        sheet_name = None
        for name in ["KSCE 실적", "Summary", "손익계산서", "PL"]:
            if name in wb.sheetnames:
                sheet_name = name
                break

        if not sheet_name:
            print(f"  [KSCE] 지원 시트 없음 — 건너뜀")
            wb.close()
            return rows

        ws = wb[sheet_name]
        year = self._detect_year(ws)
        pl_map = dict(KSCE_PL_MAP)

        # MC 데이터 존재 여부 확인
        mc_rows_exist = False
        for r in KSCE_MC_MAP:
            v = ws.cell(row=r, column=2).value
            if v and any(kw in str(v) for kw in ["재료", "노무", "경비"]):
                mc_rows_exist = True
                break

        if mc_rows_exist:
            pl_map.update(KSCE_MC_MAP)

        # 월별 데이터 추출
        for month_offset in range(12):
            col = MONTH_COL_START + month_offset
            month_num = month_offset + 1

            # 해당 월에 데이터가 있는지 확인
            test_val = ws.cell(row=4, column=col).value
            if test_val is None or safe_float(test_val) == 0:
                continue

            ym = f"{year}-{month_num:02d}"
            for row_num, (cat, sub, account) in pl_map.items():
                val = safe_float(ws.cell(row=row_num, column=col).value)
                rows.append(self.make_row(ym, "KSCE", cat, sub, account, "RSD", val))

        wb.close()
        return rows

    def _detect_year(self, ws) -> str:
        """시트에서 연도 탐지"""
        import re
        for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
            for cell in row:
                if cell:
                    m = re.search(r'(\d{4})', str(cell))
                    if m and m.group(1) in ("2024", "2025", "2026", "2027"):
                        return m.group(1)
        return "2026"
