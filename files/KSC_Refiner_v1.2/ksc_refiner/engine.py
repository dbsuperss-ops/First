import os
import sys
import json

sys.stdout.reconfigure(encoding='utf-8')
sys.stderr.reconfigure(encoding='utf-8')
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from schema import (
    MASTER_COLUMNS, load_rates, detect_company, detect_month, safe_float
)
from parsers.kscp_parser import KscpParser
from parsers.kctr_parser import KctrParser
from parsers.ksccz_parser import KscczParser
from parsers.ksce_parser import KsceParser
from parsers.ksci_parser import KsciParser

PARSER_MAP = {
    "KSCP":  KscpParser,
    "KCTR":  KctrParser,
    "KSCCZ": KscczParser,
    "KSCE":  KsceParser,
    "KSCI":  KsciParser,
}

def find_excel_files(input_dir: str, year: str = None) -> list:
    """재귀 탐색: 루트/2026/01/KCTR/*.xlsx 구조를 자동 탐지.
    단일 폴더에 파일이 있어도 동작."""
    files = []
    for root, dirs, filenames in os.walk(input_dir):
        for fn in filenames:
            if fn.startswith("~$") or "consolidated" in fn.lower():
                continue
            if fn.lower().endswith((".xlsx", ".xlsm")):
                if detect_company(fn):
                    files.append(os.path.join(root, fn))

    if year:
        year_filtered = [f for f in files if os.sep + year + os.sep in f
                         or "/" + year + "/" in f]
        if year_filtered:
            files = year_filtered

    return sorted(set(files))

def process_file(filepath: str, rates: dict) -> list:
    filename = os.path.basename(filepath)
    company = detect_company(filename)
    if not company:
        print(f"  ⚠ 법인코드 미식별: {filename}")
        return []

    parser_cls = PARSER_MAP.get(company)
    if not parser_cls:
        print(f"  ⚠ 파서 미등록: {company}")
        return []

    print(f"  📄 {filename} → {company} 파서 실행...")
    try:
        parser = parser_cls(filepath, rates)
        rows = parser.extract()
        print(f"     ✅ {len(rows)}건 추출 완료")
        return rows
    except Exception as e:
        print(f"     ❌ 파싱 실패: {e}")
        return []

def write_consolidated(all_rows: list, output_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "통합_DB"

    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="2F5496")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    for col_idx, col_name in enumerate(MASTER_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    num_fmt_local = '#,##0.00'
    num_fmt_krw = '#,##0'
    num_fmt_rate = '#,##0.00'

    for row_idx, account_row in enumerate(all_rows, 2):
        data = account_row.to_list()
        for col_idx, val in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            cell.font = Font(name="Arial", size=10)
            # MASTER_COLUMNS: 귀속연월,법인코드,데이터타입,대분류,중분류,계정과목,현지통화,현지금액,적용환율,KRW금액
            if col_idx == 8:
                cell.number_format = num_fmt_local
            elif col_idx == 9:
                cell.number_format = num_fmt_rate
            elif col_idx == 10:
                cell.number_format = num_fmt_krw

    col_widths = [12, 10, 8, 6, 10, 16, 8, 18, 10, 18]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    summary = _create_summary(wb, all_rows)
    _create_executive_report(wb, all_rows)

    wb.save(output_path)
    print(f"\n📊 통합 DB 저장: {output_path}")
    print(f"   총 {len(all_rows)}건, {len(set(r.법인코드 for r in all_rows))}개 법인")
    print(f"   포함 시트: 통합_DB, 법인별_요약, 경영보고서")

def _create_summary(wb, all_rows):
    ws = wb.create_sheet("법인별_요약")

    companies = sorted(set(r.법인코드 for r in all_rows))
    # 실적 기간만 요약 (계획 12개월이 섞이지 않도록)
    actual_rows = [r for r in all_rows if r.데이터타입 == "실적"]
    periods = sorted(set(r.귀속연월 for r in actual_rows)) or sorted(set(r.귀속연월 for r in all_rows))

    key_accounts = ["매출액", "매출원가", "매출총이익", "판관비계", "영업이익"]

    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill("solid", fgColor="2F5496")
    sub_fill = PatternFill("solid", fgColor="D6E4F0")

    headers = ["귀속연월", "계정과목"] + companies + ["합계"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    row_idx = 2
    for period in periods:
        for account in key_accounts:
            ws.cell(row=row_idx, column=1, value=period)
            ws.cell(row=row_idx, column=2, value=account)

            total = 0
            for comp_idx, comp in enumerate(companies, 3):
                val = sum(
                    r.KRW금액 for r in actual_rows
                    if r.귀속연월 == period and r.법인코드 == comp and r.계정과목 == account
                )
                ws.cell(row=row_idx, column=comp_idx, value=val).number_format = '#,##0'
                total += val

            ws.cell(row=row_idx, column=len(companies) + 3, value=total).number_format = '#,##0'
            row_idx += 1

    for i, w in enumerate([12, 14] + [16] * (len(companies) + 1), 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    ws.freeze_panes = "C2"
    return ws

def _create_executive_report(wb, all_rows):
    ws = wb.create_sheet("경영보고서")

    companies = sorted(set(r.법인코드 for r in all_rows))
    # 실적 데이터의 최신 귀속연월을 당월로 사용 (계획 12개월에 끌려가지 않도록)
    actual_rows = [r for r in all_rows if r.데이터타입 == "실적"]
    plan_rows   = [r for r in all_rows if r.데이터타입 == "계획"]
    actual_periods = sorted(set(r.귀속연월 for r in actual_rows))
    latest_period = actual_periods[-1] if actual_periods else (
        sorted(set(r.귀속연월 for r in all_rows))[-1] if all_rows else "N/A"
    )

    hdr_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    hdr_fill = PatternFill("solid", fgColor="2F5496")
    sub_hdr_fill = PatternFill("solid", fgColor="D6E4F0")
    sub_hdr_font = Font(name="Arial", bold=True, size=10)
    profit_fill = PatternFill("solid", fgColor="E2EFDA")
    loss_font = Font(name="Arial", size=10, color="FF0000")
    normal_font = Font(name="Arial", size=10)
    pct_fmt = '0.0%'
    num_fmt = '#,##0'
    thin_b = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    all_periods = sorted(set(r.귀속연월 for r in all_rows))
    period_range_start = actual_periods[0] if actual_periods else (all_periods[0] if all_periods else "N/A")

    ws.cell(row=1, column=1, value=f"경신그룹 해외법인 경영실적 종합").font = Font(name="Arial", bold=True, size=14)
    ws.cell(row=2, column=1, value=f"당월(실적): {latest_period} | 누적: {period_range_start} ~ {latest_period} | 단위: 백만원(KRW)").font = Font(name="Arial", size=10, color="808080")

    row = 4
    pl_accounts = [
        ("매출", ["매출액", "상품매출", "제품매출", "전선매출", "전장매출", "기타매출"]),
        ("매출원가", ["매출원가", "제품매출원가", "상품매출원가"]),
        ("이익", ["매출총이익"]),
        ("판관비", ["판관비계"]),
        ("이익", ["영업이익"]),
    ]
    mc_accounts = [
        ("재료비", ["재료비계"]),
        ("노무비", ["노무비계"]),
        ("경비", ["제조경비계", "기계경비"]),
    ]

    # 헤더: 구분 | 계정과목 | [법인별 실적당월] | [법인별 계획당월] | [법인별 실적누계] | [법인별 계획누계] | 합계(실적누계)
    n = len(companies)
    headers = (
        ["구분", "계정과목"]
        + [f"{c}\n실적당월" for c in companies]
        + [f"{c}\n계획당월" for c in companies]
        + [f"{c}\n실적누계" for c in companies]
        + [f"{c}\n계획누계" for c in companies]
        + ["합계\n(실적누계)"]
    )
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=ci, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_b

    row += 1

    def get_val(corp, acct, period=None, dtype=None):
        filtered = all_rows
        if dtype:
            filtered = [r for r in filtered if r.데이터타입 == dtype]
        if period:
            filtered = [r for r in filtered if r.귀속연월 == period]
        return sum(r.KRW금액 for r in filtered if r.법인코드 == corp and r.계정과목 == acct)

    def get_val_ytd(corp, acct, up_to_period, dtype=None):
        """up_to_period 이하의 누계값 반환"""
        filtered = all_rows
        if dtype:
            filtered = [r for r in filtered if r.데이터타입 == dtype]
        return sum(r.KRW금액 for r in filtered
                   if r.법인코드 == corp and r.계정과목 == acct and r.귀속연월 <= up_to_period)

    def write_account_row(ws, row, section, acct, companies, latest, is_key=False):
        ws.cell(row=row, column=1, value=section).border = thin_b
        cell_acct = ws.cell(row=row, column=2, value=acct)
        cell_acct.border = thin_b
        if is_key:
            cell_acct.font = sub_hdr_font

        total_actual_cum = 0
        for ci, comp in enumerate(companies):
            act_m  = get_val(corp=comp, acct=acct, period=latest, dtype="실적") / 1_000_000
            plan_m = get_val(corp=comp, acct=acct, period=latest, dtype="계획") / 1_000_000
            act_c  = get_val_ytd(comp, acct, latest, dtype="실적") / 1_000_000
            plan_c = get_val_ytd(comp, acct, latest, dtype="계획") / 1_000_000

            def write_cell(col, val):
                c = ws.cell(row=row, column=col, value=round(val, 0))
                c.number_format = num_fmt
                c.border = thin_b
                c.font = loss_font if val < 0 else normal_font

            write_cell(3 + ci,         act_m)   # 실적 당월
            write_cell(3 + n + ci,     plan_m)  # 계획 당월
            write_cell(3 + 2*n + ci,   act_c)   # 실적 누계
            write_cell(3 + 3*n + ci,   plan_c)  # 계획 누계

            total_actual_cum += act_c

        t_cell = ws.cell(row=row, column=3 + 4 * n, value=round(total_actual_cum, 0))
        t_cell.number_format = num_fmt
        t_cell.border = thin_b
        t_cell.font = Font(name="Arial", bold=True, size=10, color="FF0000" if total_actual_cum < 0 else "000000")

        if acct in ("매출총이익", "영업이익"):
            for c in range(1, 3 + 4 * n + 1):
                ws.cell(row=row, column=c).fill = profit_fill

    total_cols = 3 + 4 * n  # 구분+계정+실적당월n+계획당월n+실적누계n+계획누계n+합계

    ws.cell(row=row, column=1, value="【 PL 】").font = sub_hdr_font
    ws.cell(row=row, column=1).fill = sub_hdr_fill
    for ci in range(2, total_cols + 1):
        ws.cell(row=row, column=ci).fill = sub_hdr_fill
        ws.cell(row=row, column=ci).border = thin_b
    row += 1

    for section, accts in pl_accounts:
        for acct in accts:
            has_data = any(
                get_val(corp=c, acct=acct, dtype="실적") != 0 or get_val(corp=c, acct=acct, dtype="계획") != 0
                for c in companies
            )
            if has_data:
                is_key = acct in ("매출액", "매출원가", "매출총이익", "판관비계", "영업이익")
                write_account_row(ws, row, section, acct, companies, latest_period, is_key)
                row += 1

    row += 1
    ws.cell(row=row, column=1, value="【 제조원가 】").font = sub_hdr_font
    ws.cell(row=row, column=1).fill = sub_hdr_fill
    for ci in range(2, total_cols + 1):
        ws.cell(row=row, column=ci).fill = sub_hdr_fill
        ws.cell(row=row, column=ci).border = thin_b
    row += 1

    for section, accts in mc_accounts:
        for acct in accts:
            has_data = any(
                get_val(corp=c, acct=acct, dtype="실적") != 0 or get_val(corp=c, acct=acct, dtype="계획") != 0
                for c in companies
            )
            if has_data:
                write_account_row(ws, row, section, acct, companies, latest_period)
                row += 1

    row += 1
    ws.cell(row=row, column=1, value="【 이익률 (%) 】").font = sub_hdr_font
    ws.cell(row=row, column=1).fill = sub_hdr_fill
    for ci in range(2, total_cols + 1):
        ws.cell(row=row, column=ci).fill = sub_hdr_fill
    row += 1

    for label in ["매출총이익률", "영업이익률"]:
        ws.cell(row=row, column=1, value="이익률").border = thin_b
        ws.cell(row=row, column=2, value=label).border = thin_b
        ws.cell(row=row, column=2).font = sub_hdr_font

        acct = "매출총이익" if "총" in label else "영업이익"
        for ci, comp in enumerate(companies):
            # 실적 당월 이익률
            rev_m_a  = get_val(corp=comp, acct="매출액", period=latest_period, dtype="실적")
            prof_m_a = get_val(corp=comp, acct=acct,     period=latest_period, dtype="실적")
            # 계획 당월 이익률
            rev_m_p  = get_val(corp=comp, acct="매출액", period=latest_period, dtype="계획")
            prof_m_p = get_val(corp=comp, acct=acct,     period=latest_period, dtype="계획")
            # 실적 누계 이익률
            rev_c_a  = get_val_ytd(comp, "매출액", latest_period, dtype="실적")
            prof_c_a = get_val_ytd(comp, acct,     latest_period, dtype="실적")
            # 계획 누계 이익률
            rev_c_p  = get_val_ytd(comp, "매출액", latest_period, dtype="계획")
            prof_c_p = get_val_ytd(comp, acct,     latest_period, dtype="계획")

            def pct_cell(col, num, den):
                c = ws.cell(row=row, column=col, value=num / den if den else 0)
                c.number_format = pct_fmt
                c.border = thin_b

            pct_cell(3 + ci,       prof_m_a, rev_m_a)
            pct_cell(3 + n + ci,   prof_m_p, rev_m_p)
            pct_cell(3 + 2*n + ci, prof_c_a, rev_c_a)
            pct_cell(3 + 3*n + ci, prof_c_p, rev_c_p)

        row += 1

    col_widths = [10, 16] + [13] * (4 * n + 1)
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=4, column=i).column_letter].width = w

    ws.freeze_panes = "C5"
    ws.sheet_properties.tabColor = "2F5496"

def load_settings(config_dir: str = None) -> dict:
    if config_dir is None:
        # PyInstaller 실행 시 번들 리소스 경로 사용
        if getattr(sys, 'frozen', False):
            # PyInstaller로 패키징된 경우
            base_path = sys._MEIPASS
        else:
            # 일반 Python 실행 시
            base_path = os.path.dirname(__file__)
        config_dir = os.path.join(base_path, "config")

    path = os.path.join(config_dir, "settings.json")
    if os.path.exists(path):
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}

def save_settings(settings: dict, config_dir: str = None):
    if config_dir is None:
        # PyInstaller로 패키징된 경우 쓰기 불가능한 번들 경로이므로 스킵
        if getattr(sys, 'frozen', False):
            # 번들된 실행 파일에서는 설정 저장 스킵
            return
        config_dir = os.path.join(os.path.dirname(__file__), "config")

    try:
        os.makedirs(config_dir, exist_ok=True)
        path = os.path.join(config_dir, "settings.json")
        with open(path, "w", encoding="utf-8") as f:
            json.dump(settings, f, ensure_ascii=False, indent=2)
    except (OSError, PermissionError):
        # 쓰기 권한이 없는 경우 조용히 무시
        pass

def main(input_dir: str = None, output_dir: str = None, year: str = "2026"):
    # PyInstaller 실행 시 번들 리소스 경로 사용
    if getattr(sys, 'frozen', False):
        base_path = sys._MEIPASS
    else:
        base_path = os.path.dirname(os.path.abspath(__file__))

    config_dir = os.path.join(base_path, "config")
    settings = load_settings(config_dir)

    if input_dir is None:
        input_dir = settings.get("root_path", os.path.join(os.path.dirname(__file__), "..", "input"))
    if output_dir is None:
        saved_output = settings.get("output_path")
        # Windows 경로가 저장되어 있으면 무시하고 기본값 사용
        if saved_output and (saved_output.startswith("C:\\") or ":" in saved_output[:3]):
            output_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "output")
        else:
            output_dir = saved_output or os.environ.get('KSC_OUTPUT_DIR', os.path.join(os.path.dirname(os.path.abspath(__file__)), "output"))

    # 경로 정규화
    output_dir = os.path.abspath(output_dir)
    os.makedirs(output_dir, exist_ok=True)

    settings["root_path"] = input_dir
    settings["output_path"] = output_dir
    settings["last_year"] = year
    save_settings(settings, config_dir)

    print("=" * 60)
    print("🚀 KSC Settlement Refiner Engine v1.2")
    print(f"   루트 경로: {input_dir}")
    print(f"   기준 연도: {year}")
    print(f"   처리 범위: {year}년 전체 누적")
    print("=" * 60)

    rates = load_rates(year, config_dir)
    print(f"\n📌 적용 환율: {rates}")

    files = find_excel_files(input_dir, year)
    if not files:
        print("\n❌ 처리할 엑셀 파일이 없소.")
        return None

    print(f"\n📂 탐지된 파일: {len(files)}개")
    for f in files:
        print(f"   {os.path.relpath(f, input_dir)}")

    all_rows = []
    for filepath in sorted(files):
        rows = process_file(filepath, rates)
        all_rows.extend(rows)

    if not all_rows:
        print("\n❌ 추출된 데이터가 없소.")
        return None

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = os.path.join(output_dir, f"consolidated_db_{timestamp}.xlsx")
    write_consolidated(all_rows, output_path)
    return output_path

if __name__ == "__main__":
    input_dir = sys.argv[1] if len(sys.argv) > 1 else None
    year = sys.argv[2] if len(sys.argv) > 2 else "2026"
    main(input_dir=input_dir, year=year)
