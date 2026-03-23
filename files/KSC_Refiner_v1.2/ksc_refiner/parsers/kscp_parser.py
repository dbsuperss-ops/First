import openpyxl
from typing import List
from parsers import BaseParser
from schema import AccountRow, safe_float, CATEGORY_PL, CATEGORY_MC

KSCP_ACCOUNT_MAP = {
    4:  (CATEGORY_PL, "매출", "매출액"),
    6:  (CATEGORY_PL, "매출원가", "매출원가"),
    5:  (CATEGORY_PL, "이익", "매출총이익"),
    7:  (CATEGORY_PL, "이익", "영업이익"),
    8:  (CATEGORY_MC, "재료비", "재료비계"),
    9:  (CATEGORY_MC, "재료비", "원재료"),
    10: (CATEGORY_MC, "재료비", "관세"),
    11: (CATEGORY_MC, "재료비", "통관물류"),
    12: (CATEGORY_MC, "재료비", "운반비(고객)"),
    13: (CATEGORY_MC, "재료비", "재고평가손실"),
    14: (CATEGORY_MC, "노무비", "노무비계"),
    15: (CATEGORY_MC, "노무비", "직접노무비"),
    16: (CATEGORY_MC, "노무비", "간접노무비"),
    17: (CATEGORY_MC, "경비", "기계경비"),
    18: (CATEGORY_PL, "판관비", "판관비계"),
    19: (CATEGORY_PL, "판관비", "급여"),
    20: (CATEGORY_PL, "판관비", "임대용역리스"),
    21: (CATEGORY_PL, "판관비", "세금공과"),
    22: (CATEGORY_PL, "판관비", "기타판관비"),
    23: (CATEGORY_PL, "이익", "당기순이익"),
    24: (CATEGORY_PL, "영업외", "이자비용"),
}

PLAN_START_COL = 3   # C
ACTUAL_START_COL = 5  # E column = 1월

# 사업계획 파일: 사업계획 시트 KRW 손익 요약 (row 211~235, col 6=1월 ~ col 17=12월)
KSCP_PLAN_MAP = {
    211: (CATEGORY_PL, "매출", "매출액"),
    214: (CATEGORY_PL, "매출원가", "매출원가"),
    215: (CATEGORY_PL, "이익", "매출총이익"),
    217: (CATEGORY_PL, "이익", "영업이익"),
    219: (CATEGORY_MC, "재료비", "재료비계"),
    225: (CATEGORY_MC, "노무비", "노무비계"),
    229: (CATEGORY_PL, "판관비", "판관비계"),
    234: (CATEGORY_PL, "이익", "당기순이익"),
    235: (CATEGORY_PL, "영업외", "이자비용"),
}
KSCP_PLAN_MONTH_COL_START = 6  # F = 1월

class KscpParser(BaseParser):
    def extract(self) -> List[AccountRow]:
        wb = openpyxl.load_workbook(self.filepath, read_only=True, data_only=True)

        # 결산 시트 찾기: 결산 종합, 월종합, 보고용자료 등
        settlement_sheet = None
        for candidate in ["결산 종합", "월종합", "보고용자료"]:
            if candidate in wb.sheetnames:
                settlement_sheet = candidate
                break

        if not settlement_sheet:
            # 사업계획 시트 찾기
            if "사업계획" in wb.sheetnames:
                rows = self._extract_plan(wb)
            elif "KSCP 계획" in wb.sheetnames:
                rows = self._extract_plan_old(wb)
            else:
                print(f"  [KSCP] 지원 시트 없음 — 건너뜀")
                rows = []
            wb.close()
            return rows

        ws = wb[settlement_sheet]
        rows = []

        year = self._detect_year(wb)

        for month_offset in range(12):
            actual_col = ACTUAL_START_COL + month_offset
            month_num = month_offset + 1

            test_val = ws.cell(row=4, column=actual_col).value
            if test_val is None or safe_float(test_val) == 0:
                continue

            ym = f"{year}-{month_num:02d}"

            for row_num, (cat, sub, account) in KSCP_ACCOUNT_MAP.items():
                val = safe_float(ws.cell(row=row_num, column=actual_col).value)
                rows.append(self.make_row(ym, "KSCP", cat, sub, account, "KRW", val))

        wb.close()
        return rows

    def _extract_plan(self, wb) -> List[AccountRow]:
        """사업계획 시트 KRW 손익 요약 (rows 211~235) 에서 월별 데이터 추출"""
        ws = wb["사업계획"]
        year = self._detect_year_plan(ws)
        rows = []
        for month_offset in range(12):
            col = KSCP_PLAN_MONTH_COL_START + month_offset
            month_num = month_offset + 1
            test_val = ws.cell(row=211, column=col).value
            if test_val is None or safe_float(test_val) == 0:
                continue
            ym = f"{year}-{month_num:02d}"
            for row_num, (cat, sub, account) in KSCP_PLAN_MAP.items():
                val = safe_float(ws.cell(row=row_num, column=col).value)
                rows.append(self.make_row(ym, "KSCP", cat, sub, account, "KRW", val, data_type="계획"))
        return rows

    def _extract_plan_old(self, wb) -> List[AccountRow]:
        """KSCP 계획 시트 (구 형식) 에서 월별 데이터 추출"""
        ws = wb["KSCP 계획"]
        year = self._detect_year_plan(ws)
        rows = []
        for month_offset in range(12):
            col = KSCP_PLAN_MONTH_COL_START + month_offset
            month_num = month_offset + 1
            test_val = ws.cell(row=211, column=col).value
            if test_val is None or safe_float(test_val) == 0:
                continue
            ym = f"{year}-{month_num:02d}"
            for row_num, (cat, sub, account) in KSCP_PLAN_MAP.items():
                val = safe_float(ws.cell(row=row_num, column=col).value)
                rows.append(self.make_row(ym, "KSCP", cat, sub, account, "KRW", val, data_type="계획"))
        return rows

    def _detect_year_plan(self, ws) -> str:
        import re
        for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
            for cell in row:
                if cell:
                    m = re.search(r'(\d{4})', str(cell))
                    if m and m.group(1) in ("2024", "2025", "2026", "2027"):
                        return m.group(1)
        return "2026"

    def _detect_year(self, wb) -> str:
        for sheet_name in ["결산 종합", "KSCP양식", "KSCP 계획"]:
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
                for cell in row:
                    if cell and "2026" in str(cell):
                        return "2026"
                    if cell and "2025" in str(cell):
                        return "2025"
        return "2026"
