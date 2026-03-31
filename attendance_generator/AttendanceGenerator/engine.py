"""
출석부 자동 생성 엔진 v1.1
- 맞춤형: 날짜 5개까지 지원 (G/H/J/L/N 컬럼)
- 방과후: 월별 파일 분리 생성 (양식 최대 6칸)
Usage: python engine.py <명단파일.xlsx> <시작일YYYY-MM-DD> <종료일YYYY-MM-DD> <출력폴더>
"""
import sys
import os
import json
from datetime import date, timedelta
from copy import copy
from openpyxl import load_workbook, Workbook

DAY_KO_MAP = {'월': 0, '화': 1, '수': 2, '목': 3, '금': 4, '토': 5, '일': 6}
CUSTOM_DATE_COLS     = [7, 8, 10, 12, 14]      # 맞춤형: 최대 5개
AFTERSCHOOL_DATE_COLS = [8, 9, 11, 13, 15, 17] # 방과후: 최대 6개


def get_weekdays_in_range(start, end, weekday):
    days, d = [], start
    while d <= end:
        if d.weekday() == weekday:
            days.append(d)
        d += timedelta(days=1)
    return days


def group_by_month(dates):
    groups = {}
    for d in dates:
        groups.setdefault((d.year, d.month), []).append(d)
    return groups


def load_roster(filepath):
    wb = load_workbook(filepath, read_only=True)
    ws = wb['명단']
    rows = list(ws.iter_rows(values_only=True))
    data = []
    for row in rows[7:]:
        구분 = row[1]
        if not 구분 or str(구분).strip() in ('아침돌봄', '오후돌봄'):
            continue
        data.append({
            '구분':     str(구분).strip(),
            '과목':     str(row[2]).strip() if row[2] else '',
            '요일':     str(row[3]).strip() if row[3] else '',
            '강사':     str(row[5]).strip() if row[5] else '',
            '이름':     str(row[7]).strip() if row[7] else '',
            '학년':     row[8] if row[8] else '',
            '반':       row[9] if row[9] else '',
            '번호':     row[10] if row[10] else '',
            '전화번호': str(row[11]).strip() if row[11] else '',
            '참여요일': str(row[14]).strip() if row[14] else '',
            '비고':     str(row[15]).strip() if row[15] else '',
        })
    wb.close()
    return data


def get_combinations(roster):
    seen, result = set(), []
    for r in roster:
        key = (r['구분'], r['과목'], r['요일'])
        if key not in seen:
            seen.add(key)
            result.append(key)
    return sorted(result)


def copy_sheet_to_new_wb(template_wb, sheet_name):
    import copy as cp
    src = template_wb[sheet_name]
    new_wb = Workbook()
    dst = new_wb.active
    dst.title = sheet_name

    for col_letter, cd in src.column_dimensions.items():
        dst.column_dimensions[col_letter].width = cd.width
    for row_idx, rd in src.row_dimensions.items():
        dst.row_dimensions[row_idx].height = rd.height

    for row in src.iter_rows():
        for cell in row:
            dst_cell = dst.cell(row=cell.row, column=cell.column)
            if cell.value is not None:
                dst_cell.value = cell.value
            if cell.has_style:
                dst_cell.font          = cp.copy(cell.font)
                dst_cell.border        = cp.copy(cell.border)
                dst_cell.fill          = cp.copy(cell.fill)
                dst_cell.number_format = cell.number_format
                dst_cell.protection    = cp.copy(cell.protection)
                dst_cell.alignment     = cp.copy(cell.alignment)

    for mc in src.merged_cells.ranges:
        dst.merge_cells(str(mc))

    dst.page_setup.paperSize   = src.page_setup.paperSize
    dst.page_setup.orientation = src.page_setup.orientation
    dst.page_margins = cp.copy(src.page_margins)
    return new_wb, dst


def 학년_label(students):
    years = set(str(s['학년']) for s in students)
    if years == {'1'}: return '1학년'
    if years == {'2'}: return '2학년'
    return '1·2학년'


def fill_맞춤형(template_wb, students, 과목, 요일, start_date, end_date, output_path):
    if not students:
        return False
    weekday_num = DAY_KO_MAP.get(요일, -1)
    if weekday_num < 0:
        return False
    dates = get_weekdays_in_range(start_date, end_date, weekday_num)
    if not dates:
        return False

    new_wb, ws = copy_sheet_to_new_wb(template_wb, '맞춤형 출석부')
    강사 = students[0]['강사']

    ws.cell(3, 6).value = f'초1~2 맞춤형 {과목}({요일}) 프로그램 {학년_label(students)} 출석부'
    ws.cell(5, 4).value = f'{start_date.year}. {start_date.month}. {start_date.day}.~{end_date.year}. {end_date.month}. {end_date.day}.'
    ws.cell(6, 9).value = f'{강사}  (인)'

    for i, d in enumerate(dates[:len(CUSTOM_DATE_COLS)]):
        col = CUSTOM_DATE_COLS[i]
        ws.cell(8, col).value = d
        ws.cell(8, col).number_format = 'M/D'

    for idx, s in enumerate(students):
        r = 10 + idx
        ws.cell(r, 2).value  = idx + 1
        ws.cell(r, 3).value  = s['이름']
        ws.cell(r, 5).value  = s['반']
        ws.cell(r, 6).value  = s['참여요일']
        ws.cell(r, 16).value = s['비고']
        ws.cell(r, 19).value = s['전화번호']

    new_wb.save(output_path)
    return True


def fill_방과후_single(template_wb, students, 과목, 요일, dates,
                       month_start, month_end, output_path):
    if not students or not dates:
        return False

    new_wb, ws = copy_sheet_to_new_wb(template_wb, '방과후 출석부')
    강사 = students[0]['강사']

    ws.cell(2, 6).value  = f' {과목}({학년_label(students)}) 1학기 출석부'
    ws.cell(4, 4).value  = f'{month_start.year}. {month_start.month}. {month_start.day}.~\n{month_end.year}. {month_end.month}. {month_end.day}.'
    ws.cell(4, 10).value = f'{요일}요일 강사'
    ws.cell(5, 10).value = f'{강사}  (인)'

    for i, d in enumerate(dates[:len(AFTERSCHOOL_DATE_COLS)]):
        col = AFTERSCHOOL_DATE_COLS[i]
        ws.cell(7, col).value = d
        ws.cell(7, col).number_format = 'M/D'

    for idx, s in enumerate(students):
        r = 9 + idx
        ws.cell(r, 2).value  = idx + 1
        ws.cell(r, 3).value  = s['이름']
        ws.cell(r, 5).value  = s['학년']
        ws.cell(r, 6).value  = s['반']
        ws.cell(r, 7).value  = s['번호']
        ws.cell(r, 17).value = s['비고']
        ws.cell(r, 20).value = s['전화번호']

    new_wb.save(output_path)
    return True


def fill_방과후(template_wb, students, 구분, 과목, 요일, start_date, end_date, output_dir):
    weekday_num = DAY_KO_MAP.get(요일, -1)
    if weekday_num < 0:
        return []

    all_dates = get_weekdays_in_range(start_date, end_date, weekday_num)
    if not all_dates:
        return []

    MONTH_KO = ['','1월','2월','3월','4월','5월','6월',
                '7월','8월','9월','10월','11월','12월']
    created = []
    for (year, month), dates in group_by_month(all_dates).items():
        safe = f'{구분}_{과목}_{요일}_{MONTH_KO[month]}_출석부.xlsx'
        safe = safe.replace(' ', '_').replace('/', '_')
        path = os.path.join(output_dir, safe)
        ok = fill_방과후_single(template_wb, students, 과목, 요일,
                               dates, dates[0], dates[-1], path)
        if ok:
            created.append(os.path.basename(path))
    return created


def generate_all(roster_path, start_str, end_str, output_dir):
    start_date = date.fromisoformat(start_str)
    end_date   = date.fromisoformat(end_str)
    os.makedirs(output_dir, exist_ok=True)

    roster      = load_roster(roster_path)
    combos      = get_combinations(roster)
    template_wb = load_workbook(roster_path)

    for (구분, 과목, 요일) in combos:
        students = [r for r in roster
                    if r['구분']==구분 and r['과목']==과목 and r['요일']==요일]
        try:
            if 구분 == '맞춤형':
                safe = f'{구분}_{과목}_{요일}'.replace(' ','_').replace('/','_')
                path = os.path.join(output_dir, f'{safe}_출석부.xlsx')
                ok   = fill_맞춤형(template_wb, students, 과목, 요일,
                                   start_date, end_date, path)
                result = {'combo': f'{구분} / {과목} / {요일}',
                          'files': [os.path.basename(path)] if ok else [],
                          'status': 'ok' if ok else 'skip'}
            else:
                files  = fill_방과후(template_wb, students, 구분, 과목, 요일,
                                    start_date, end_date, output_dir)
                result = {'combo': f'{구분} / {과목} / {요일}',
                          'files': files,
                          'status': 'ok' if files else 'skip'}
        except Exception as e:
            result = {'combo': f'{구분} / {과목} / {요일}',
                      'files': [], 'status': f'error: {e}'}

        print(json.dumps(result, ensure_ascii=False), flush=True)


if __name__ == '__main__':
    if len(sys.argv) < 5:
        print('Usage: python engine.py <명단.xlsx> <시작일> <종료일> <출력폴더>')
        sys.exit(1)
    generate_all(sys.argv[1], sys.argv[2], sys.argv[3], sys.argv[4])
